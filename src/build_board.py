"""
Builds the Excel draft board from data/player_features.csv.

WHY THIS FILE EXISTS
--------------------
Through v7 the board was assembled ad hoc, which made it the only
artifact in the project with no source. That's backwards -- the board is
the actual deliverable, and everything upstream exists to produce it.
It also meant the ranking logic that matters most on draft day (VOR,
replacement levels, draft targets) lived nowhere it could be reviewed or
re-run.

Every rule below was reverse-engineered from 2026_Draft_Board_v7.xlsx and
verified against it: replacement levels, VOR, sort order, value delta,
and draft-target text all reproduce v7 given v7's inputs.

USAGE
-----
    # 12-team league -- writes 2026_12Team_Board_v9.xlsx
    python -m src.build_board

    # 6-team league -- writes 2026_6Team_Board_v9.xlsx
    python -m src.build_board --config league_config_6team.json

    python -m src.build_board --note "refreshed ADP"
    python -m src.build_board --output some/path.xlsx

ONE FILE PER LEAGUE
-------------------
The filename carries the league and the MODEL version, and nothing else.
Rebuilding overwrites in place and appends a row to the in-workbook "Build
History" sheet, so the count of rebuilds is visible without accumulating a
folder of near-identical spreadsheets. Bump MODEL_VERSION when weights or
features change; a data refresh (new ADP, edited injury overrides) keeps the
same version and is told apart by the git hash and timestamp in that sheet.

Requires openpyxl (not currently in .venv):
    pip install openpyxl
"""

import argparse
import json
import math
import subprocess
from datetime import datetime
from pathlib import Path

import polars as pl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.ranking import load_situational_weights
from src.vona_columns import snake_gaps, default_slot, compute_ppg_pos_rank
from src.playing_time import (
    expected_games_for_rookies,
    load_playing_time_model,
)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
FEATURES_PATH = PROJECT_ROOT / "data" / "player_features.csv"
CONFIG_PATH = PROJECT_ROOT / "league_config_12team.json"

# MODEL_VERSION bumps when the model changes -- new weights, new features,
# a refit. It does NOT bump for a data refresh (new ADP pull, updated injury
# overrides), because nothing about the ranking logic moved. Rebuilds of the
# same model version are told apart by the Build History sheet, not by the
# filename, which is what keeps one file per league instead of a pile of
# near-identical spreadsheets.
# Phase 10 (Aug 4): 9 -> 10. This one is an honest bump, unlike the 8 -> 9
# documented in PHASE_8-14_PLAN.md. Features changed (age replaced experience;
# usage_trend_share added at RB and TE) and all three positions were refit,
# so ranks move for real.
# Phase 11 (Aug 4): 10 -> 11. No refit -- the weights are byte-identical to
# v10. Replacement level changed from starter slots to expected players
# drafted (CP6), which moves every VOR and reorders the board, so it is a
# ranking-logic change and bumps the version even though no coefficient did.
# Phase 11 A+B (Aug 4): 11 -> 12. The baseline definition itself changed
# twice -- `discount_thin` season weighting (CP3) and shrinkage toward the
# position's 30th percentile at K=2 (CP5) -- and all four positions were
# refit against the new deltas. RB `trend_missing` drops out of the model
# entirely as a result. This is the largest single move since Phase 7.
# Phase 12 (Aug 6): 12 -> 13. Rookies stop taking a zero adjustment and
# start taking a fitted one at RB and WR, so rookie ranks move on EVERY
# board, not only the new one. That is a ranking-logic change by this
# constant's own definition and has to bump it.
#
# It nearly didn't. Phase 12 was built alongside a new league, and the
# attention was on that league -- but `ranking.apply_situational_weights`
# is shared, so the 12-team and 6-team boards both changed the moment
# rookie_weights.json appeared. A board whose numbers moved while its
# version stayed put is the exact problem the Build History sheet was
# added to solve, one level up: there, three rebuilds of one version; here,
# two models under one version. Rebuild all three boards off this bump.
#
# 13 -> 14 (Aug 12), Phase 13.5. Rookie availability now feeds
# `expected_games`, so Exp Gm and Exp Pts change on all three boards. No
# rank moves -- the phase's gate chose availability alone and left the
# rate untouched -- but "the numbers moved" is the bump rule and Exp Pts
# is a number. A version that only tracked rank would be a version that
# lies about the column this phase exists to fix.
#
# 14 -> 15 (Aug 12), Phase 13.5b. Availability refitted on a LOGIT scale.
# Every rookie's Exp Gm changes, most of them slightly and undrafted
# quarterbacks a great deal -- pick 245 reads 4.1% instead of a clamped
# 0.0%. Same rule as the 13->14 bump: the numbers moved. Rank still does
# not, and `compare_boards --expect rank-identical` against v14 is the
# check.
#
# 15 -> 16 (Aug 12), Phase 13.6. Two changes on the 32-team board, both
# ranking logic by this constant's own definition. Replacement level moves
# (QB59/RB92/WR145/TE56 -> QB62/RB99/WR162/TE61, the average of two mocks
# on a 12-round scale), which reprices every VOR. And ADP itself changes
# source, from the FFC 12-team 2QB feed to the mocks -- which reorders far
# more of the sheet than the replacement move does, because `has_adp`
# gates ABOVE vor in the sort and ~120 players cross that gate.
#
# The other two boards are untouched: `apply_mock_adp` is a no-op without
# `adp_mock_file` in the config, and their replacement levels are derived
# as before. They are rebuilt at v16 anyway, so that three boards sitting
# in one folder on draft day can't be from two different versions.
#
# 16 -> 17 (Aug 12), Phase 13.7. `position_overrides.csv` adds Travis Hunter
# to the universe on all three boards.
#
# The bump was predicted on the wrong grounds and the measurement is worth
# keeping. The claim was that adding a player to a position pool moves other
# players: he joins the WR group `position_competition_top_k` averages over
# at JAX and the pool baseline shrinkage anchors against. MEASURED against
# v16: zero Jacksonville players moved, and 2 of 450 receivers -- both for
# unrelated reasons. His PPG sits below the top-K competition set at his own
# team, and one player in a 450-man pool does not shift a percentile anchor.
# Every rank below his 166 shifted by exactly +1, which is an insertion, not
# a revaluation.
#
# It still bumps, on the honest reason rather than the guessed one: the
# board contains a player it did not contain before. A version that only
# tracked whether existing numbers moved would call two boards with
# different rosters the same version.
MODEL_VERSION = 17

# Last row of the merged notes block. The header row sits two below it.
# Phase 14 moved this from a hardcoded 6 (with the header hardcoded at 8)
# because the notes tripled in length and the two numbers have to move
# together or the instructions get clipped by the header.
NOTES_LAST_ROW = 10

HISTORY_SHEET = "Build History"
INJURY_OVERRIDES_PATH = PROJECT_ROOT / "injury_overrides.csv"

# Status in injury_overrides.csv that removes a player from contention.
# Any OTHER status string is recorded in the Notes column but changes no
# ranking -- so you can jot "QUESTIONABLE - hamstring" as a reminder
# without it silently moving anyone.
OUT_STATUS = "OUT_SEASON"

# Phase 11 CP8. Statuses that cost a player a KNOWN, PARTIAL slice of the
# season rather than all of it. PUP and NFI are not opinions about health --
# they are roster designations with a rule attached: a player who opens the
# regular season on either list is ineligible for the first four games. That
# is a floor, not an estimate, which is what separates these from
# QUESTIONABLE and lets them move a number at all.
#
# Anything longer is player-specific and belongs in the optional
# `games_missed` column of injury_overrides.csv, which overrides this
# default. Kittle's Achilles is the obvious candidate -- four games is
# almost certainly generous to him.
PARTIAL_STATUSES = {"PUP", "NFI"}
PUP_DEFAULT_GAMES_MISSED = 4

# How the single FLEX slot is expected to be filled across the league.
# This is a modeling ASSUMPTION, not a league rule -- it's the standard
# full-PPR split and it only moves replacement levels by a fraction of a
# roster spot. Documented here because it silently shifts every VOR.
#
# Phase 11 CP6 demoted this to a fallback: it is now used only by
# compute_starter_ranks(), which no longer drives VOR. See
# compute_replacement_ranks() for why.
FLEX_SPLIT = {"RB": 0.40, "WR": 0.40, "TE": 0.20}

# Same idea for a SUPERFLEX slot, which is FLEX plus quarterbacks.
#
# The split is overwhelmingly QB and that is not a modeling preference,
# it is arithmetic: in any league where a second quarterback is legal,
# the worst startable QB outscores the best startable flex RB/WR by more
# than any other swap available, so superflex slots go to quarterbacks
# until the position runs dry. The residual is the tail of teams who
# missed the QB run and start a third receiver there.
#
# Like FLEX_SPLIT this is an ASSUMPTION, it is only used by the legacy
# compute_starter_ranks(), and it does not drive VOR. It exists so that
# the starter-count floor in compute_replacement_ranks() knows a
# superflex league starts more than `num_teams` quarterbacks -- without
# it the floor is num_teams x 0, and QB replacement level would be free
# to land absurdly shallow on a board where QB is the scarcest position
# there is.
SUPERFLEX_SPLIT = {"QB": 0.85, "RB": 0.05, "WR": 0.08, "TE": 0.02}

# Roster slots every team fills but this model does not rank. They still
# consume real picks, so they must come out of the pick pool before skill
# positions are allocated -- otherwise a 16-round draft looks like it has
# 2 x num_teams more skill picks than it does, and replacement level lands
# too deep at every position.
#
# WAS A HARDCODED 2 UNTIL THE 32-TEAM LEAGUE (Aug 6)
# --------------------------------------------------
# Both leagues that existed when this was written started 1 K and 1 DST,
# so `2` was correct twice and looked like a constant. The 32-team
# superflex league starts neither, and the constant would have removed
# 64 picks from a 320-pick draft that does not spend them -- pulling
# replacement level ~20% shallower at every position and inflating every
# VOR on the board. It fails silently and in the direction that makes
# the board look more confident, which is the worst combination.
#
# Derived from roster_slots now, so a league that adds or drops a
# non-skill slot gets the right pick pool without a second place to
# remember.
UNMODELED_SLOTS = ("K", "DST")

MODELED_POSITIONS = ["QB", "RB", "WR", "TE"]

# How many of the deepest ADP-covered picks define the mix used to
# extrapolate a draft longer than the feed. See compute_replacement_ranks.
#
# 40 is a compromise with a reason on each side. Too narrow and the mix
# is noise -- a single positional run at the end of the feed would
# distort the whole extrapolation. Too wide and it stops being a TAIL
# and becomes the overall mix again, which is the behaviour being
# replaced. 40 is roughly the last four rounds of a 10-team draft: long
# enough to average over a run, short enough to still describe late-round
# behaviour.
TAIL_WINDOW = 40

# Extrapolating this much of a draft is not a footnote. At or above this
# share the run says so in plain language rather than printing a NOTE
# that reads like every other NOTE.
EXTRAPOLATION_WARN_SHARE = 0.25


def unmodeled_slots_per_team(config):
    """Starting slots per team that consume a pick but get no ranking."""
    slots = config["roster_slots"]
    return sum(slots.get(name, 0) for name in UNMODELED_SLOTS)

FONT_NAME = "Arial"

# Row fill by position, for players inside the draftable pool.
POSITION_FILLS = {
    "QB": "CFE2F3",  # light blue
    "RB": "D9EAD3",  # light green
    "WR": "FCE5CD",  # light orange
    "TE": "EAD1DC",  # light mauve
}
# Phase 11 B (CP4). Baseline confidence, shaded onto the GP (sample) cell
# rather than given a column of its own. The board was reordered for
# density and this is a warning, not a number you read -- you want to
# notice it while looking at something else.
#
# Thresholds are games in the 3-year window: under one full season is
# amber, under half a season is red. Shrinkage has already corrected the
# projection by the time you see this; the shading says how much of what
# you are looking at is the model's prior rather than the player's record.
LOW_CONFIDENCE_GAMES = 17
VERY_LOW_CONFIDENCE_GAMES = 8
LOW_CONFIDENCE_FILL = "FFE599"       # amber
VERY_LOW_CONFIDENCE_FILL = "F4B183"  # orange

UNDRAFTABLE_FILL = "D9D9D9"  # gray: has ADP but ranks past the last pick
NO_ADP_FILL = "F4CCCC"       # pink: no real ADP, hard-capped to the bottom
OUT_FILL = "E06666"          # strong red: out for the season, do not draft
HEADER_FILL = "1F4E78"
INJURY_FILL = "FF0000"

# Ordered by what you actually reach for on the clock, left to right, in
# four blocks. Through v11 the order was an accident of when each column
# got added, which put "Has ADP" -- a filter helper rendered in invisible
# text -- eight columns left of the bye week.
#
#   1. WHO      Rank / Pos / Player, frozen so they survive scrolling.
#   2. ACT      Draft Target answers the only question the clock asks.
#               Then VOR (what he's worth), Adj PPG (the projection under
#               it), Value Δ and the two ADP columns (what the room thinks).
#   3. CHECK    Why, then the things that veto a pick you'd otherwise make:
#               bye collisions, availability, injury.
#   4. AUDIT    Model internals. Real, but not draft-day reading.
COLUMNS = [
    # -- 1. WHO --------------------------------------------------------
    ("Rank", 7), ("Pos", 6), ("Player", 24),
    # -- 2. ACT --------------------------------------------------------
    # Phase 14 reordered this block. Through v17 it opened with Draft
    # Target, on the reasoning that it "answers the only question the
    # clock asks." The draft simulator retired that claim: Draft Target is
    # ADP minus one standard deviation, a per-player heuristic that does
    # not know when your next pick is, and VONA answers the same question
    # against your actual pick schedule. Keeping the weaker version in the
    # first column invites you to run two conflicting rules on the clock,
    # so it moves to CHECK and the inputs VONA actually consumes move up.
    #
    # What VONA needs, in the order you touch it: the projection, where
    # the room takes him, and his rank among his own position -- because
    # VONA asks a POSITIONAL question and this sheet is sorted globally.
    #
    # REMOVED AGAIN IN THE SAME PHASE: "Wait 5" / "Wait 19", which
    # precomputed the cost of passing on a player. They were built, tested
    # and cut, and the cut is the more useful result. Two reasons:
    #
    #   1. They do not work well enough. A policy that drafted straight off
    #      the frozen columns scored +18.8 against best-available where the
    #      same policy computing the survivor LIVE scored +59.7. Freezing
    #      the ADP assumption throws away two thirds of the edge, because
    #      the correction that matters -- "four backs just went in six
    #      picks, so count four" -- is exactly the one a precomputed column
    #      cannot make.
    #   2. They are not comparable across rows. The number is anchored to
    #      each player's own ADP, so a tail player is measured against the
    #      people behind HIM. Joe Mixon at ADP 187 outscored Jahmyr Gibbs,
    #      which reads as "bigger cliff" and means "cliff nobody will fall
    #      off." A column you have to remember not to sort is a trap.
    #
    # The live calculation off Adj PPG + PPG@Pos + ADP (Ovr) is three
    # columns and one subtraction, and it is worth three times as much.
    # compute_wait_cost() survives in src/vona_columns.py because
    # src/drift_test.py is what measured all of the above.
    ("Adj PPG", 9), ("ADP (Ovr)", 10), ("PPG@Pos", 9),
    ("VOR", 8), ("Bye", 6),
    # -- 3. CHECK ------------------------------------------------------
    # Draft Target and Value Δ are now reference rather than instruction;
    # see the notes block on the sheet.
    ("Draft Target", 24), ("Value Δ (picks)", 12), ("ADP (Rd.Pk)", 12),
    # Phase 11. Plain-language decomposition of Sit Adj: the largest
    # signed contributions, relative to an average player at that
    # position. Generated from the same weights that produce the number
    # in Sit Adj, so the two cannot disagree.
    ("Why (value drivers)", 46),
    ("Team", 7), ("Recent Injury", 12),
    # Phase 11 CP8. Availability, kept OUT of the ranking on purpose.
    # Exp Gm is this league's regular season minus known PUP/NFI absence;
    # Exp Pts is Adj PPG x Exp Gm. Adj PPG stays an honest per-game rate,
    # so Kittle does not move in the Rank column -- the cost of missing
    # September shows up here instead, and it is genuinely a different
    # number in each league (12-game regular season in the 6-team
    # league, 14 in the 12-team), which no single rank column could express.
    ("Exp Gm", 8), ("Exp Pts", 9),
    # -- 4. AUDIT ------------------------------------------------------
    ("Sit Adj", 8), ("Rook", 6),
    # Phase 10. Age is a model input at RB/WR/TE (it replaced
    # `experience`, which it beat at every position). Usage Trend is
    # also a model input at all three -- WR only after the training set
    # widened from three seasons to five, which took it from p=0.23 to
    # p=0.034.
    #
    # "Trend n" is how many seasons the slope was fitted on. 3 is a full
    # slope, 2 is a two-point slope, 0 means none could be fitted and the
    # player was mean-imputed. Worth showing rather than hiding: the
    # trend signal turned out to be CARRIED by the n=2 players, not
    # weakened by them.
    ("Age", 6), ("Usage Trend", 11), ("Trend n", 8), ("GP (sample)", 11),
    ("Has ADP", 4),
    ("Notes (manual)", 45),
]

# Column label -> 1-based sheet index. Phase 10 left a comment warning that
# "if you insert another column, these all move again" above a block of
# hardcoded indices; Phase 11 inserts three columns, so that warning gets
# retired rather than obeyed. Look columns up by name.
COLUMN_INDEX = {label: i for i, (label, _) in enumerate(COLUMNS, start=1)}


def load_config(path=CONFIG_PATH):
    with open(path) as f:
        return json.load(f)


def league_slug(config):
    """
    The filename stem. Resolution order: `board_label`, then `config_key`,
    then the league name.

    `board_label` exists because a board is a function of RULES, not of a
    league's name. Jack runs several 12-team leagues on identical
    settings, and one "12Team" board serves all of them -- three
    byte-identical files named after three commissioners would be worse
    than one named after what it actually is.

    THE HAZARD THIS CREATES, stated plainly: two configs sharing a
    `board_label` write to the SAME file, and the second build silently
    overwrites the first. That is exactly what you want while the rules
    match and exactly what you don't the moment they diverge. So if you
    ever add a 12-team league with different scoring, a different keeper
    rule, or a different roster, give it its own label. The label being an
    explicit field rather than something derived from `num_teams` is what
    makes that a decision you make rather than a collision you discover on
    draft day.
    """
    # `board_label` ships VERBATIM. str.capitalize() lowercases everything
    # after the first character, which turned "12Team" into "12team" --
    # the label is already written the way it should appear, so nothing
    # should be reformatting it. `config_key` is snake_case by convention
    # and still gets converted.
    label = config.get("board_label")
    if label:
        return str(label).strip().replace(" ", "")

    key = config.get("config_key")
    if key:
        return "".join(part.capitalize() for part in str(key).split("_"))

    words = [w for w in str(config["league_name"]).split() if w.lower() != "fantasy"]
    return "".join(w.capitalize() for w in words[:2])


def git_short_hash():
    """
    Short commit hash, or 'nogit' if unavailable.

    This is what actually distinguishes two rebuilds of the same model
    version, so it's worth having even when it fails softly.
    """
    try:
        result = subprocess.run(
            ["git", "-C", str(PROJECT_ROOT), "rev-parse", "--short", "HEAD"],
            capture_output=True, text=True, timeout=5, check=True,
        )
        return result.stdout.strip() or "nogit"
    except Exception:
        return "nogit"


def read_build_history(output_path):
    """
    Pulls the existing Build History rows out of a board we're about to
    overwrite.

    This is the mechanism that lets one file per league still answer "how many
    times has this been rebuilt." The alternative -- a new file per build --
    answers the same question by accumulating near-identical spreadsheets, and
    then you have to remember which one you printed.

    A build history that silently resets to 1 would be worse than none at all,
    so a corrupt or unreadable existing file raises rather than starting over.
    """
    output_path = Path(output_path)
    if not output_path.exists():
        return []
    workbook = load_workbook(output_path, read_only=True)
    if HISTORY_SHEET not in workbook.sheetnames:
        workbook.close()
        return []
    sheet = workbook[HISTORY_SHEET]
    rows = [
        [cell for cell in row]
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row and row[0] is not None
    ]
    workbook.close()
    return rows


def apply_injury_overrides(players, path=INJURY_OVERRIDES_PATH):
    """
    Joins the manually-maintained injury_overrides.csv onto the player
    table, adding `out_for_season` (bool) and `injury_note` (str).

    WHY THIS IS MANUAL
    ------------------
    The model's own `recent_major_injury` feature only flags players who
    ended the 2025 REGULAR SEASON on IR. It is structurally blind to
    anything that happens in the 2026 offseason or preseason, and no
    nflverse feed reliably covers August injuries -- official injury
    reports don't begin until Week 1.

    That blindness is dangerous specifically BECAUSE ADP gets refreshed.
    The market drops an injured player within hours; the model's
    projection doesn't move at all, since it's built from trailing
    2023-25 per-game rates. `value_delta` is (adp_rank - model_rank), so
    the collapse in ADP shows up as a huge positive value delta and the
    board recommends him as a bargain. Ricky Pearsall went from -11 to
    roughly +87 in exactly this way. Without this file, every injury
    between now and draft day makes the board worse, not staler.

    File format -- player_name,status,note[,games_missed]
    Only status == OUT_STATUS changes the RANKING. Anything else is
    carried into the Notes column and leaves rank alone.

    PHASE 11 CP8 -- PUP AND NFI
    ---------------------------
    A third case sits between "fine" and "gone." PUP and NFI are not
    health opinions, they are roster designations with a rule attached: a
    player who opens the regular season on either list cannot play the
    first four games. Treating that as a note (the old behavior) had
    Kittle and Charbonnet showing at full value; treating it as
    OUT_SEASON would be a lie in the other direction.

    So those statuses set `expected_games_missed` -- four by default, or
    whatever the optional `games_missed` column says for a specific
    player. That number feeds Exp Gm / Exp Pts and NOTHING else. It does
    not touch `adjusted_fantasy_points_per_game`, so it cannot move rank
    or VOR.

    That restraint is deliberate and it is the phase's open question,
    answered: PPG is a RATE, and a torn Achilles doesn't make Kittle worse
    per game he plays -- it makes him play fewer of them. Folding
    availability into the rate would corrupt the one number the whole
    model is fitted to predict, and would silently double-count the moment
    a future phase models availability directly. Season-long value lives
    in its own column instead, where it can be read against the rate
    rather than baked into it.

    UNMATCHED NAMES RAISE. A typo'd name silently matching nothing would
    leave an injured player sitting in the draftable pool looking like a
    steal -- the exact failure this file exists to prevent -- so it fails
    loudly instead.
    """
    empty = players.with_columns([
        pl.lit(False).alias("out_for_season"),
        pl.lit(None, dtype=pl.String).alias("injury_note"),
        pl.lit(0.0).alias("expected_games_missed"),
        pl.lit(None, dtype=pl.String).alias("injury_status"),
    ])

    if not Path(path).exists():
        print(f"No {Path(path).name} found -- no injury overrides applied.")
        return empty

    overrides = pl.read_csv(path)
    if overrides.height == 0:
        print(f"{Path(path).name} is empty -- no injury overrides applied.")
        return empty

    known_names = set(players.select("player_name").to_series().to_list())
    unmatched = [
        name for name in overrides.select("player_name").to_series().to_list()
        if name not in known_names
    ]
    if unmatched:
        raise ValueError(
            f"{Path(path).name}: these names match no player in "
            f"player_features.csv: {unmatched}. Fix the spelling to match "
            f"the nflverse name exactly -- an unmatched override does "
            f"nothing, which would leave an injured player draftable."
        )

    status = (pl.col("status").cast(pl.String).str.to_uppercase()
              .str.strip_chars())

    # `games_missed` is optional -- the file is hand-maintained and most
    # rows have no reason to carry one. Absent column, absent value, and
    # blank string all fall back to the PUP default.
    if "games_missed" in overrides.columns:
        stated_games = pl.col("games_missed").cast(pl.Float64, strict=False)
    else:
        stated_games = pl.lit(None, dtype=pl.Float64)

    overrides = overrides.select([
        "player_name",
        (status == OUT_STATUS).alias("out_for_season"),
        pl.col("note").cast(pl.String).alias("injury_note"),
        status.alias("injury_status"),
        pl.when(status.is_in(list(PARTIAL_STATUSES)))
          .then(stated_games.fill_null(float(PUP_DEFAULT_GAMES_MISSED)))
          .otherwise(0.0)
          .alias("expected_games_missed"),
    ])

    joined = players.join(overrides, on="player_name", how="left").with_columns([
        pl.col("out_for_season").fill_null(False),
        pl.col("expected_games_missed").fill_null(0.0),
    ])

    out_count = joined.select(pl.col("out_for_season").sum()).item()
    partial = joined.filter(pl.col("expected_games_missed") > 0)
    print(f"Injury overrides: {overrides.height} entries, {out_count} marked {OUT_STATUS}, "
          f"{partial.height} on PUP/NFI")
    for row in partial.select(["player_name", "injury_status",
                               "expected_games_missed"]).iter_rows():
        print(f"   {row[0]}: {row[1]}, -{row[2]:.0f} games")
    return joined


def compute_expected_points(players, config):
    """
    Phase 11 CP8. Adds `expected_games` and `expected_total_points`.

    `expected_games` is this LEAGUE'S regular season minus known absence,
    and the denominator is the regular season on purpose -- not the
    17-week NFL calendar. Weeks 15-17 are worth nothing in a league whose
    final is week 14, and nothing at all in the 6-team league, whose season ends in
    week 12. Those are the games that decide whether you reach the
    playoffs at all.

    That is why the same four-game PUP absence costs 4/12 = 33% of a
    6-team season against 4/14 = 29% of a 12-team one, and why the
    same player is honestly worth different amounts in the two leagues --
    something the board has never had to express before.

    Out-for-season players get zero expected games, which is the one place
    Exp Pts and the ranking agree.
    """
    season = float(config.get("fantasy_season_length", 17))

    expected_games = (
        pl.when(pl.col("out_for_season"))
        .then(pl.lit(0.0))
        .otherwise(
            (pl.lit(season) - pl.col("expected_games_missed")).clip(0.0, season)
        )
    )

    return players.with_columns([
        expected_games.alias("expected_games"),
        (pl.col("adjusted_fantasy_points_per_game") * expected_games)
        .alias("expected_total_points"),
    ])


# Phase 11. How each model feature reads in English on the board.
#
# `label` receives the player's raw feature value so a driver can carry the
# number that caused it ("age 31" beats "age"), and the SIGN of the printed
# contribution always comes from the arithmetic, never from the wording --
# so a feature whose coefficient flips in a later refit relabels itself
# correctly with no edit here.
#
# Note the two mean-reversion terms. `workload_share` and
# `position_competition_ppg` both carry negative weights, which reads
# backwards until you remember what the regression predicts: not PPG, but
# the DELTA from a player's own trailing baseline. A back already holding
# 60% of his backfield has his usage priced in and little room left to
# grow; a receiver whose position group is otherwise weak has already
# banked that. So the wording describes the situation, not a judgment.
DRIVER_LABELS = {
    "age": lambda v: f"age {v:.0f}",
    "workload_share": lambda v: f"{v * 100:.0f}% team share",
    "usage_trend_share": lambda v: f"role trend {v * 100:+.1f}pp/yr",
    "trend_missing": lambda v: "thin usage history",
    "qb_changed": lambda v: "new QB" if v else "same QB",
    "team_changed": lambda v: "new team" if v else "same team",
    "recent_major_injury": lambda v: "2025 IR" if v else "no 2025 IR",
    "position_competition_ppg": lambda v: f"teammates {v:.1f} PPG",
    "experience": lambda v: f"{v:.0f} yrs exp",
    "continuity_score": lambda v: f"continuity {v:.2f}",
}

# A driver has to move the projection by at least this much to be worth a
# reader's attention mid-draft. Below it the term is real but not
# actionable, and printing four of them buries the one that matters.
DRIVER_MIN_ABS = 0.15
DRIVER_MAX_TERMS = 4


def build_value_drivers(players, weights_by_position=None):
    """
    Phase 11. Adds `value_drivers`: a short signed list of what is pushing
    a player's projection up or down, largest effect first.

        "-1.4 62% team share · +0.9 role trend +3.1pp/yr · -0.5 age 30"

    Each term is a real number out of the fitted model, not a description
    of one:

        contribution_f = (player_value_f - position_mean_f) x weight_f

    so the terms are deviations from an AVERAGE player at that position,
    and they sum -- with the position's base offset -- exactly to the Sit
    Adj column beside them. That identity is the whole point. A "why"
    column written by hand, or generated from a separate set of rules,
    would eventually contradict the number it is explaining, and on draft
    day you would have no way to tell which one was lying. This one cannot
    drift: it is computed from the same JSON that computes the adjustment,
    and verify_adjustments.py asserts the two reconcile.

    Two housekeeping notes. Nulls are filled with the position mean, which
    is what apply_situational_weights() does, so an imputed feature
    contributes exactly 0 and correctly never appears as a driver -- "no
    opinion" should not read as a reason. And the position base offset
    (intercept plus the mean-value terms) is around -0.1 PPG at every
    position, small enough to omit from the string unless it clears the
    same threshold as everything else.

    Rookies take no situational adjustment at all, so they say so rather
    than showing an empty cell that looks like missing data.
    """
    if weights_by_position is None:
        weights_by_position = load_situational_weights()

    rows = players.to_dicts()
    drivers = []

    for row in rows:
        if row.get("is_rookie"):
            drivers.append("rookie — cohort baseline, no situational adj")
            continue

        spec = weights_by_position.get(row["position"])
        if not spec:
            # K/DST never reach here, and QB did until Phase 10.
            drivers.append(None)
            continue

        means = spec.get("feature_means", {})
        centers = spec.get("centers", {})

        terms = []
        base = float(spec["intercept"])
        for feature, weight in spec["weights"].items():
            mean = float(means.get(feature, 0.0))
            center = float(centers.get(feature, 0.0))

            # Same fill rule as apply_situational_weights(): a missing
            # value means "average player," not zero.
            raw = row.get(feature)
            value = mean if raw is None else float(raw)

            base += (mean - center) * weight
            contribution = (value - mean) * weight
            if abs(contribution) >= DRIVER_MIN_ABS:
                label = DRIVER_LABELS.get(feature, lambda v, f=feature: f)(value)
                terms.append((abs(contribution), f"{contribution:+.1f} {label}"))

        if abs(base) >= DRIVER_MIN_ABS:
            terms.append((abs(base), f"{base:+.1f} position base"))

        terms.sort(key=lambda t: -t[0])
        text = " · ".join(t[1] for t in terms[:DRIVER_MAX_TERMS])
        if not text:
            text = "average situation — no driver above ±0.1"

        # Availability is not part of Sit Adj and must not look like it
        # is, so it goes after a pipe rather than into the sum.
        # Phase 11 B (CP4). If shrinkage moved this player materially,
        # say so and by how much. A projection that is largely the
        # model's prior rather than the player's record is the single
        # most important thing to know about it, and it belongs next to
        # the other reasons rather than inferred from a shaded cell.
        games = row.get("games_played")
        shrink = row.get("baseline_shrink_delta")
        if games is not None and shrink is not None and abs(shrink) >= DRIVER_MIN_ABS:
            text += f" | {shrink:+.1f} thin sample ({games:.0f} gm)"

        missed = row.get("expected_games_missed") or 0.0
        if row.get("out_for_season"):
            text += " | OUT for season"
        elif missed > 0:
            text += f" | {row.get('injury_status') or 'INJ'} −{missed:.0f} gm"

        drivers.append(text)

    return players.with_columns(pl.Series("value_drivers", drivers, dtype=pl.String))


GATE_PATH = PROJECT_ROOT / "data" / "holdout_gate.json"
WEIGHT_FILES = [
    PROJECT_ROOT / "data" / "situational_weights.json",
    PROJECT_ROOT / "data" / "rookie_weights.json",
]


def require_holdout_gate(skip=False):
    """
    Refuses to build a board unless the current weights have passed
    out-of-sample validation.

    WHY THIS IS A HARD STOP (Aug 6). Phase 13 CP2 cut four things that
    alpha had passed and that leave-one-season-out had also passed --
    including a whole position's only feature, and a phase's headline
    finding. In-sample significance has now demonstrably admitted
    features that do not predict. The gate is the only check downstream
    of a refit that would catch the next one.

    THREE WAYS TO FAIL, all of them things that would otherwise happen
    silently:

      MISSING   No gate has ever been run against these weights.
      FAILED    The gate ran and something shipped has no out-of-sample
                value.
      STALE     The weights are NEWER than the gate. This is the
                dangerous one and the reason mtimes are compared rather
                than just reading `passed`: refitting after a green gate
                leaves a passing file on disk describing a model that no
                longer exists. Exactly the failure the Build History
                sheet was added to solve, and the same shape as the
                stale-weights check already in verify_adjustments.

    `--skip-gate` exists for genuine emergencies and says so loudly. It
    is not for "the gate is annoying."
    """
    if skip:
        print("\n  *** HOLDOUT GATE SKIPPED (--skip-gate) ***")
        print("  This board may contain features that do not predict out of "
              "sample. Do not draft from it without knowing why you skipped.\n")
        return

    if not GATE_PATH.exists():
        raise SystemExit(
            f"\nBUILD BLOCKED: no holdout gate at {GATE_PATH.name}.\n"
            f"These weights have never been validated out of sample, which is how "
            f"four non-predictive features reached this model once already.\n\n"
            f"Run:  python -m src.holdout --gate\n"
        )

    with open(GATE_PATH) as f:
        gate = json.load(f)

    gate_time = GATE_PATH.stat().st_mtime
    stale = [p.name for p in WEIGHT_FILES
             if p.exists() and p.stat().st_mtime > gate_time]
    if stale:
        raise SystemExit(
            f"\nBUILD BLOCKED: {stale} are NEWER than the holdout gate.\n"
            f"The gate on disk passed a model that has since been refitted, so it "
            f"describes weights that no longer exist. A green light for the wrong "
            f"model is worse than no light.\n\n"
            f"Run:  python -m src.holdout --gate\n"
        )

    if not gate.get("passed"):
        failures = "\n".join(f"    {x}" for x in gate.get("failures", []))
        raise SystemExit(
            f"\nBUILD BLOCKED: the holdout gate failed.\n\n{failures}\n\n"
            f"Remove these from FEATURE_SPECS, refit, and re-run the gate. "
            f"Shipping them means drafting off features that are known not to "
            f"predict.\n"
        )

    print(f"Holdout gate: PASSED (folds {gate.get('seasons')})")
    require_playing_time_gate(skip=skip)


PLAYING_TIME_PATH = PROJECT_ROOT / "data" / "playing_time.json"
PLAYING_TIME_GATE_PATH = PROJECT_ROOT / "data" / "playing_time_gate.json"


def require_playing_time_gate(skip=False):
    """
    The same three checks, for the Phase 13.5 model.

    WHY A SECOND FUNCTION RATHER THAN TWO MORE ENTRIES IN `WEIGHT_FILES`.
    The staleness rule is "the artifact must not be newer than the gate
    that passed it," and each model has its OWN gate. Adding
    playing_time.json to WEIGHT_FILES would compare it against
    holdout_gate.json -- a gate that never tested it. That is precisely
    the "green light for the wrong model" this file already refuses to
    accept, so it gets its own pairing rather than borrowing one.

    ABSENT IS ALLOWED HERE, and that is the one real difference. A board
    with no playing-time model is v13: rookies keep a full season of
    expected games, which is wrong but is what shipped for months and
    does not affect a single rank. A board with STALE or FAILED playing
    time is a board asserting a correction it cannot support, and that
    is blocked like any other.
    """
    if skip or not PLAYING_TIME_PATH.exists():
        return

    if not PLAYING_TIME_GATE_PATH.exists():
        raise SystemExit(
            f"\nBUILD BLOCKED: {PLAYING_TIME_PATH.name} exists but "
            f"{PLAYING_TIME_GATE_PATH.name} does not.\n"
            f"The rookie availability model has never been validated out of "
            f"sample.\n\n"
            f"Run:  python -m src.playing_time --gate\n"
        )

    if PLAYING_TIME_PATH.stat().st_mtime > PLAYING_TIME_GATE_PATH.stat().st_mtime:
        raise SystemExit(
            f"\nBUILD BLOCKED: {PLAYING_TIME_PATH.name} is NEWER than its gate.\n"
            f"The playing-time model has been refitted since it was last "
            f"validated.\n\n"
            f"Run:  python -m src.playing_time --gate\n"
        )

    with open(PLAYING_TIME_GATE_PATH) as f:
        pt_gate = json.load(f)

    if not pt_gate.get("passed"):
        raise SystemExit(
            f"\nBUILD BLOCKED: the playing-time gate failed.\n"
            f"Delete {PLAYING_TIME_PATH.name} to build without rookie "
            f"availability, or fix the model.\n"
        )

    ship = pt_gate.get("ship", "?")
    print(f"Playing-time gate: PASSED (ships predictor {ship}, "
          f"n={pt_gate.get('n_held_out')})")


def rescore_for_league(players, config, base_config_path=CONFIG_PATH):
    """
    Re-expresses every projection in THIS league's scoring.

    THE BUG THIS FIXES (Aug 6)
    --------------------------
    `fantasy_points_per_game` is computed once, in features.py, under one
    config -- league_config_12team.json. That was invisible while
    both leagues used identical scoring blocks. They did: the 12-team and
    6-team configs differ in teams, weeks, and keepers, and in nothing
    that touches a point value.

    The 32-team league is the first with different SCORING: 4-point
    passing touchdowns and -1 interceptions against 6 and -2. Without
    this function its board would rank quarterbacks on numbers roughly
    2.5-3.5 PPG too high apiece -- a passing touchdown is worth a third
    less and every one of them was counted at full price. In a SUPERFLEX
    league, where QB is the scarcest position on the board and 27 of them
    get drafted, that is not a rounding error; it is the single largest
    distortion the board could carry.

    Worth noting how quietly it fails. Every number still looks
    plausible, the columns all populate, and nothing anywhere raises.
    The board just tells you to draft quarterbacks.

    HOW
    ---
    Only `pass_td` and `interception` can differ between the configs
    this project ships, and player_features.csv already stores
    `passing_tds_per_game` and `passing_interceptions_per_game`. So the
    correction is exact, not modelled:

        delta = pass_tds_pg  x (league.pass_td   - base.pass_td)
              + pass_ints_pg x (league.interception - base.interception)

    Any OTHER scoring key differing is caught and raised rather than
    silently ignored, because the moment a league changes `reception` or
    `rush_td` this arithmetic is no longer complete and a wrong answer
    would be worse than a crash.

    TWO THINGS THIS DOES NOT FIX, both stated rather than papered over:

    1. ROOKIE QUARTERBACKS. A rookie's projection is a cohort baseline,
       not a stat line, so he has no `passing_tds_per_game` to correct
       with -- the columns are null. Rookie QBs therefore stay in the
       BASE league's scoring on this board and are overvalued by roughly
       the same 2.5-3.5 PPG. The count is small (a handful clear the
       depth chart) but on a superflex board they sit exactly where it
       hurts. Fixing it properly means scoring the cohort baselines per
       league in rookies.py, which is a pipeline change, not a board one.
       Flagged loudly at the bottom of this function.

    2. THE FITTED WEIGHTS. Situational coefficients were fitted against
       deltas measured in base scoring. QB carries exactly one weight
       (`age`, -0.19 PPG/yr) and its level shift is already suppressed,
       so the residual error is a fraction of a point and does not
       reorder anything. Recorded because it is real, not because it is
       urgent -- a league that changed `reception` would make this
       matter at every position at once, and that league is the one that
       should refit.
    """
    with open(base_config_path) as f:
        base_scoring = json.load(f)["scoring"]
    league_scoring = config["scoring"]

    correctable = {"pass_td", "interception"}
    differing = {
        k for k in league_scoring
        if isinstance(league_scoring.get(k), (int, float))
        and league_scoring.get(k) != base_scoring.get(k)
    }

    uncorrectable = differing - correctable
    if uncorrectable:
        raise ValueError(
            f"{config['league_name']} differs from the base scoring config in "
            f"{sorted(uncorrectable)}, which rescore_for_league() cannot correct "
            f"from the columns in player_features.csv.\n"
            f"Every projection on this board would be in the WRONG SCORING and "
            f"nothing downstream would notice.\n"
            f"Fix by making features.py score per league, or extend this function "
            f"with the per-game columns those keys need. Do not remove this check."
        )

    if not differing:
        return players

    td_delta = league_scoring["pass_td"] - base_scoring["pass_td"]
    int_delta = league_scoring["interception"] - base_scoring["interception"]

    correction = (
        pl.col("passing_tds_per_game").cast(pl.Float64).fill_null(0.0) * td_delta
        + pl.col("passing_interceptions_per_game").cast(pl.Float64).fill_null(0.0)
        * int_delta
    )

    ppg_columns = [
        c for c in [
            "fantasy_points_per_game",
            "fantasy_points_per_game_shrunk",
            "adjusted_fantasy_points_per_game",
        ] if c in players.columns
    ]
    players = players.with_columns(
        [(pl.col(c) + correction).alias(c) for c in ppg_columns]
    )

    moved = players.filter(correction.abs() >= 0.5)
    print(f"Rescored for {config['league_name']}: pass_td {base_scoring['pass_td']}"
          f"->{league_scoring['pass_td']}, int {base_scoring['interception']}"
          f"->{league_scoring['interception']} "
          f"({moved.height} players moved 0.5+ PPG across {len(ppg_columns)} columns)")

    # Anyone the correction could not reach, for whatever reason. Keyed
    # on the RATE BEING NULL rather than on `is_rookie`, which is the
    # honest test: since Aug 6 rookies carry their cohort's average
    # passing rates (see rookies.aggregate_rookie_season), so being a
    # rookie no longer implies being uncorrectable. Checking the thing
    # that actually blocks the arithmetic means this warning stays true
    # if the data changes underneath it.
    stranded = players.filter(
        (pl.col("position") == "QB") & pl.col("passing_tds_per_game").is_null()
    )
    if stranded.height:
        # Scoped to the ones that could actually reach a roster. The raw
        # count is dominated by undrafted camp bodies who will never be
        # picked in any league, and a warning that cries wolf about 21
        # players when 2 matter is a warning that gets ignored on draft
        # day. `has_adp` is the honest filter: it means a real drafter,
        # somewhere, took this player on purpose.
        draftable = stranded.filter(pl.col("has_adp").cast(pl.String)
                                    .str.to_lowercase().eq("true"))
        print(f"  NOTE: {stranded.height} QB(s) have no passing rate to rescore "
              f"from and stay in the base scoring.")
        if draftable.height:
            names = ", ".join(
                draftable.sort("adp", nulls_last=True)
                .select("player_name").to_series().to_list()[:5]
            )
            print(f"  WARNING: {draftable.height} of them have an ADP and are "
                  f"OVERVALUED on this board by ~2-3 PPG: {names}")
            print(f"  They remain in {base_scoring['pass_td']}-point-passing-TD "
                  f"scoring. Discount them by hand. See rescore_for_league().")
        else:
            print(f"  None of them carry an ADP, so none is realistically "
                  f"draftable and the distortion does not reach the board.")

    return players


def select_adp_variant(players, config):
    """
    Points the canonical `adp` / `has_adp` columns at whichever ADP feed
    this league's config asks for.

    pipeline.py attaches every variant in adp.ADP_VARIANTS to one
    player_features.csv -- `ppr` unsuffixed and `2qb` as `adp_2qb` --
    because all boards must ship from a single model run. A league with
    `"adp_format": "2qb"` gets those suffixed columns renamed over the
    canonical ones HERE, once, so that every function downstream
    (compute_replacement_ranks, compute_draft_targets, the sheet writer)
    goes on reading `adp` and never learns there was a choice.

    Renaming beats threading a column name through fifteen call sites,
    and it beats a config-aware `adp` in the CSV, which would make
    player_features.csv league-specific and quietly break Phase 13 CP4's
    one-model-many-boards rule.
    """
    adp_format = config.get("adp_format", "ppr")
    if adp_format == "ppr":
        return players

    suffix = f"_{adp_format}"
    suffixed = [c for c in players.columns if c.endswith(suffix)]
    if not suffixed:
        print(
            f"\n  WARNING: {config['league_name']} asks for adp_format="
            f"{adp_format!r} but player_features.csv has no *{suffix} columns.\n"
            f"  Falling back to the default PPR feed, which is NOT what this "
            f"league drafts under. Re-run `python -m src.pipeline` to pull it.\n"
        )
        return players

    # Drop the canonical columns first, then rename the suffixed ones onto
    # the names they just vacated.
    canonical = [c[: -len(suffix)] for c in suffixed]
    players = players.drop([c for c in canonical if c in players.columns])
    players = players.rename(dict(zip(suffixed, canonical)))

    print(f"ADP: using the {adp_format!r} feed for {config['league_name']} "
          f"({len(suffixed)} columns remapped)")
    return players


def apply_mock_adp(players, config):
    """
    PHASE 13.6. Replaces the FFC feed with ADP measured in this league's own
    mock drafts, for any config carrying an `adp_mock_file`.

    WHY A LEAGUE WOULD WANT THIS
    ----------------------------
    FFC's deepest 2026 2QB data stops around pick 190. A 32-team draft is 384
    picks long, so from round 7 onward every remaining player had `has_adp`
    false: hard-capped below every ADP-bearing player and shaded pink, in the
    exact half of the draft the board exists to help with. Two real 32-team
    mocks cover all 12 rounds, and they describe a 32-team superflex room
    rather than a 12-team 2QB room rescaled and hoped over.

    IT REPLACES THE FEED RATHER THAN FILLING ITS GAPS
    -------------------------------------------------
    The tempting version keeps FFC where it exists and uses the mocks only
    past pick 190. That would be wrong in a way nothing downstream could
    catch: pick 150 in a 12-team draft and pick 150 in a 32-team draft are
    different players' worth of draft capital, and compute_replacement_ranks
    reads the ADP ORDER of the first `skill_picks` names to decide where
    replacement level sits at every position. Mixing two scales inside one
    ordering silently corrupts that. One scale, or the other.

    The cost is stated rather than hidden: two mocks is a much thinner sample
    than FFC's hundreds, and one of them had most teams autodrafting. The
    `times_drafted` and `adp_stdev` columns carry that thinness onto the sheet
    -- a player taken in only one mock shows stdev in the dozens of picks,
    which is what widens his Draft Target cushion instead of pretending to a
    precision two drafts cannot support.
    """
    mock_file = config.get("adp_mock_file")
    if not mock_file:
        return players

    path = PROJECT_ROOT / mock_file
    if not path.exists():
        raise FileNotFoundError(
            f"{config['league_name']} asks for mock ADP at {mock_file}, which "
            f"does not exist. Build it with `python -m src.mock_adp`, or drop "
            f"`adp_mock_file` from the config to fall back to the FFC feed."
        )

    mock = pl.read_csv(path).select([
        "player_id",
        pl.col("adp").alias("mock_adp"),
        pl.col("adp_high").alias("mock_adp_high"),
        pl.col("adp_low").alias("mock_adp_low"),
        pl.col("adp_stdev").alias("mock_adp_stdev"),
        pl.col("times_drafted").alias("mock_times_drafted"),
    ])

    players = players.join(mock, on="player_id", how="left")
    players = players.with_columns([
        pl.col("mock_adp").alias("adp"),
        pl.col("mock_adp_high").alias("adp_high"),
        pl.col("mock_adp_low").alias("adp_low"),
        pl.col("mock_adp_stdev").alias("adp_stdev"),
        pl.col("mock_times_drafted").alias("times_drafted"),
        pl.col("mock_adp").is_not_null().alias("has_adp"),
    ]).drop([c for c in players.columns if c.startswith("mock_")])

    covered = players.filter(pl.col("has_adp")).height
    once = players.filter(pl.col("times_drafted") == 1).height
    print(f"ADP: {config['league_name']} reads MOCK DRAFT ADP from {mock_file} "
          f"-- {covered} players covered, {once} of them from a single mock. "
          f"The FFC feed is not used on this board.")
    return players


def league_gaps(config):
    """
    The two snake-draft gaps for this league, from `draft_slot` in the
    config. Falls back to the middle of the room when the slot is unknown,
    which is the least-wrong single guess: it is the only slot whose two
    gaps are both close to the average.
    """
    teams = int(config["num_teams"])
    slot = int(config.get("draft_slot") or default_slot(teams))
    slot = max(1, min(teams, slot))
    return snake_gaps(teams, slot), slot


def add_vona_columns(board, config):
    """
    Phase 14. Attaches `ppg_pos_rank` -- rank within position by projected
    points, best = 1.

    One column, and it is the one that survived. The board is sorted by
    VOR, which is a GLOBAL order; the question you ask on every pick is
    positional ("who is the best tight end left"). Without this you filter
    and scan the sheet on the clock. With it you read down a column.

    A DISPLAY column. It does not touch `adjusted_fantasy_points_per_game`,
    VOR, or the sort, so no rank moves and MODEL_VERSION does not bump. The
    model did not change; the way you read it did.
    """
    rows = board.to_dicts()
    return board.with_columns(
        pl.Series("ppg_pos_rank", compute_ppg_pos_rank(rows), dtype=pl.Int32)
    )


def compute_starter_ranks(config):
    """
    LEGACY (Phase 8 - Phase 10). How many players at each position get
    drafted as STARTERS league-wide.

    Superseded by compute_replacement_ranks() below, and retained only so
    build_notes() can show both numbers -- the gap between them is the
    single biggest change in v11 and hiding it would make the board's
    movement inexplicable.
    """
    teams = config["num_teams"]
    slots = config["roster_slots"]
    flex_slots = slots.get("FLEX", 0)
    superflex_slots = slots.get("SUPERFLEX", 0)

    ranks = {}
    for position in MODELED_POSITIONS:
        starters = teams * slots.get(position, 0)
        flex_share = teams * flex_slots * FLEX_SPLIT.get(position, 0.0)
        superflex_share = teams * superflex_slots * SUPERFLEX_SPLIT.get(position, 0.0)
        ranks[position] = max(1, round(starters + flex_share + superflex_share))
    return ranks


def compute_replacement_ranks(config, players):
    """
    Phase 11 CP6. How many players at each position come off the board
    across the WHOLE draft. The next one down is replacement level: the
    best player still sitting on waivers when the draft ends.

    WHY THE OLD RULE WAS WRONG
    --------------------------
    Replacement used to be the last STARTER -- QB12 in a 12-team league,
    which is defensible, but QB6 in a 6-team one, which is not. In a
    6-team league QB7 through QB32 are all unowned, so the quarterback you
    can have for nothing is a fine starting NFL quarterback, not the worst
    rostered one. Setting replacement at QB6 measured every quarterback
    against a bar almost nobody has to clear, and the board responded by
    putting Josh Allen 5th overall in a league where you can stream the
    position. The bug was invisible at 12 teams, where starter count and
    waiver depth roughly agree; it only surfaced when a second league
    forced the comparison.

    The fix is to count picks, not starters. A 6-team 16-round draft is 96
    picks, of which 12 go to kickers and defenses, leaving 84 for skill
    positions -- against 168 in the 12-team league. Fewer picks means a
    shallower cut into every position, which is the actual mechanism by
    which a shallow league makes waivers rich.

    HOW THE 84 GET SPLIT
    --------------------
    By observed draft behavior: the position mix of the first N players in
    ADP order, out-for-season players removed since nobody spends a pick
    on them. This is the plan's "expected players drafted per position"
    and it is the one place ADP earns its keep -- it is the only evidence
    the project has about what drafters actually do, and it enters through
    replacement level rather than through any player's projection, so the
    statistics-only rule on `adjusted_fantasy_points_per_game` holds.

    KNOWN LIMITATION, stated plainly: FFC's ADP comes from 12-team mocks,
    so the mix of its first 84 picks is a 12-team drafter's mix, not a
    6-team drafter's. A 6-team room, facing no scarcity at all, would
    almost certainly take FEWER than the 8 quarterbacks this yields. The
    error therefore runs in the conservative direction -- it understates
    how far QB should fall -- and a league can override the split outright
    with an `expected_drafted` block in its config if you ever have real
    draft results to fit.

    Sanity condition from the plan: the 6-team board must push QB and TE
    DOWN relative to the 12-team board. It does. Josh Allen goes from 7th
    to 15th on the 6-team board while rising to 7th on the 12-team, and the
    quarterbacks inside the top 30 go 4 -> 1 on the shallow board and
    1 -> 4 on the deep one.
    """
    teams = config["num_teams"]
    rounds = config["total_rounds"]
    skill_picks = teams * rounds - teams * unmodeled_slots_per_team(config)

    override = config.get("expected_drafted")
    if override:
        ranks = {p: int(override[p]) for p in MODELED_POSITIONS if p in override}
        total = sum(ranks.values())

        # A HAND-ENTERED OVERRIDE THAT DOES NOT SUM IS SILENTLY WRONG
        # (Aug 6). These counts come off a real draft board, typed in by
        # a human, and they replace the ADP machinery entirely -- so
        # nothing downstream can notice if they are short or long. Ten
        # missing picks would move replacement level at every position
        # and shift every VOR on the board with no error anywhere.
        #
        # The counts describe a draft, and a draft has a known length.
        # That makes this checkable, so it gets checked.
        if total != skill_picks:
            raise ValueError(
                f"`expected_drafted` in {config['league_name']} sums to {total}, "
                f"but this draft has {skill_picks} skill picks "
                f"({teams} teams x {rounds} rounds"
                + (f" minus {teams * unmodeled_slots_per_team(config)} for K/DST"
                   if unmodeled_slots_per_team(config) else "")
                + f").\n"
                f"Counts: {ranks}\n"
                f"A miscount here moves replacement level at every position and "
                f"nothing downstream would notice. Fix the counts, or the roster "
                f"if the league changed."
            )

        print(f"Replacement from OBSERVED draft results ({total} picks): {ranks}")
        return ranks

    # SUPERFLEX / ADP-FORMAT NOTE (Aug 6). The position mix below is read
    # off whichever ADP feed select_adp_variant() installed, and for a
    # superflex league the feed matters more than anything else in this
    # function -- it is the sole evidence for how many quarterbacks come
    # off the board, and QB is the position superflex changes.
    #
    #   adp_format="ppr"  -> one-QB mocks. Understates QB demand badly.
    #                        The starter floor is doing all the work.
    #   adp_format="2qb"  -> the closest feed FFC publishes, and the right
    #                        call. Still not identical: 2QB REQUIRES two
    #                        starters where superflex only permits a
    #                        second, so it leans slightly the other way.
    #
    # Either way the number is a proxy, and which direction it leans is
    # worth printing next to it rather than leaving in a comment nobody
    # opens on draft day.
    if config["roster_slots"].get("SUPERFLEX", 0) and not config.get("expected_drafted"):
        adp_format = config.get("adp_format", "ppr")
        floor = compute_starter_ranks(config).get("QB")
        if adp_format == "ppr":
            print(
                f"\n  WARNING: {config['league_name']} has a SUPERFLEX slot but reads "
                f"the non-superflex 'ppr' ADP feed.\n"
                f"  The derived QB count will UNDERSTATE how many quarterbacks go, "
                f"pushing QB replacement too shallow and UNDERVALUING every "
                f"quarterback here. The starter floor (QB{floor}) is the only thing "
                f"holding it up.\n"
                f"  Set \"adp_format\": \"2qb\" in the config.\n"
            )
        else:
            print(
                f"\n  NOTE: {config['league_name']} is superflex and reads the "
                f"{adp_format!r} ADP feed -- the closest proxy FFC publishes, but a "
                f"proxy. 2QB requires a second starting QB where superflex only "
                f"permits one, so this OVERSTATES QB demand a little.\n"
                f"  Starter floor is QB{floor}. Override with an `expected_drafted` "
                f"block once real draft data exists.\n"
            )

    # Draft order = ADP order. Out-for-season players are dropped: they
    # consume no pick, so leaving them in would push replacement one slot
    # shallower at their position for no reason.
    drafted = (
        players
        .filter(pl.col("has_adp") & ~pl.col("out_for_season"))
        .sort("adp", nulls_last=True)
        .head(skill_picks)
    )

    counts = dict(
        drafted.group_by("position").len().iter_rows()
    )
    ranks = {p: int(counts.get(p, 0)) for p in MODELED_POSITIONS}

    # If the ADP feed is shorter than the draft, the uncovered picks have
    # to be attributed to positions somehow, or replacement lands too
    # shallow and every VOR on the board inflates.
    #
    # UNIFORM SCALING WAS WRONG, AND THE 32-TEAM BOARD IS WHERE IT SHOWED
    # -------------------------------------------------------------------
    # The old rule multiplied every position's count by
    # skill_picks/observed. That assumes the position mix of the covered
    # part of the draft continues unchanged through the uncovered part.
    # It is harmless when the shortfall is small and catastrophic when it
    # is not.
    #
    # First run of the 32-team superflex board: ADP covered 184 of 320
    # picks, so 42% of the draft was extrapolated at 1.74x, and QB
    # replacement came out at **QB63**. There are 32 starting
    # quarterbacks in the NFL. QB63 is a third-stringer projecting near
    # zero, which handed every real quarterback a VOR roughly equal to
    # his entire projection and pushed Josh Allen to 3rd overall.
    #
    # The flaw is that QB and TE draw from a FINITE startable pool while
    # RB and WR do not. Rounds 8-10 of any draft are running back and
    # receiver dart throws; nobody is taking a 40th quarterback. Scaling
    # QB by the same factor as WR asserts otherwise.
    #
    # THE FIX USES THE FEED'S OWN TAIL
    # --------------------------------
    # Extrapolate the shortfall using the position mix of the DEEPEST
    # covered picks rather than the mix of the whole feed. If
    # quarterbacks have stopped going by the time the feed runs out, the
    # tail mix says so and the extrapolation stops adding them -- with no
    # invented parameter and no hand-set cap. It is still drafter
    # behaviour, still the one place ADP earns its keep, just read at the
    # point that actually describes the picks being estimated.
    observed = sum(ranks.values())
    if observed and observed < skill_picks:
        shortfall = skill_picks - observed
        tail = drafted.tail(min(TAIL_WINDOW, drafted.height))
        tail_counts = dict(tail.group_by("position").len().iter_rows())
        tail_total = sum(tail_counts.get(p, 0) for p in MODELED_POSITIONS)

        if tail_total:
            share = {p: tail_counts.get(p, 0) / tail_total for p in MODELED_POSITIONS}
        else:  # pragma: no cover -- only if the tail has no modeled positions
            share = {p: ranks[p] / observed for p in MODELED_POSITIONS}

        print(f"NOTE: ADP covers {observed} of {skill_picks} skill picks. "
              f"Attributing the remaining {shortfall} using the mix of the last "
              f"{tail.height} covered picks:")
        print("      " + "  ".join(
            f"{p} {share[p]:.0%}" for p in MODELED_POSITIONS
        ))
        ranks = {
            p: max(1, ranks[p] + round(shortfall * share[p]))
            for p in MODELED_POSITIONS
        }

        if shortfall / skill_picks > EXTRAPOLATION_WARN_SHARE:
            print(f"      WARNING: {shortfall / skill_picks:.0%} of this draft is "
                  f"extrapolated, not observed. Replacement level here is an "
                  f"estimate resting on {tail.height} picks.")
            print(f"      Set an `expected_drafted` block in the config as soon as "
                  f"you have real draft results.")

    # A position must have at least as many drafted as there are starting
    # slots; otherwise replacement lands above a player somebody has to
    # start. Only bites on a feed with almost no ADP data.
    floors = compute_starter_ranks(config)
    ranks = {p: max(ranks[p], floors.get(p, 1)) for p in ranks}

    return ranks


def compute_vor(players, replacement_ranks, value_column="adjusted_fantasy_points_per_game"):
    """
    Value Over Replacement: a player's PPG minus his position's
    replacement-level PPG.

    This is what makes the board draftable. Ranking by raw PPG would put
    22 quarterbacks at the top -- Josh Allen scores more than any running
    back alive -- but in a 1-QB league the relevant question isn't "who
    scores most," it's "how much do I gain over the guy I could get for
    free at this position." Allen is worth ~5 points over QB12; Puka
    Nacua is worth ~9 over WR29. That's the real gap.
    """
    frames = []
    for position, rank in replacement_ranks.items():
        pool = players.filter(pl.col("position") == position).sort(
            value_column, descending=True, nulls_last=True
        )
        if pool.height == 0:
            continue

        # Players who are out for the season don't count toward
        # replacement level -- nobody can start them, so they aren't the
        # freely-available alternative this whole calculation is about.
        # Removing one from above the replacement rank pulls a slightly
        # worse player into that slot, which lowers replacement PPG and
        # correctly makes everyone else at the position a bit more
        # valuable. VOR is still computed FOR them, so the board can show
        # what they'd have been worth healthy.
        available = pool.filter(~pl.col("out_for_season"))
        if available.height == 0:
            available = pool

        # If a position has fewer players than its replacement rank,
        # fall back to the worst available rather than indexing off the end.
        index = min(rank, available.height) - 1
        replacement_ppg = available.select(value_column).to_series()[index]
        frames.append(
            pool.with_columns(
                (pl.col(value_column) - replacement_ppg).alias("vor")
            )
        )
    return pl.concat(frames, how="vertical")


def format_slot(overall_pick, teams):
    """
    Turns an overall pick number into 'Beginning/Middle/End of Round N'.

    Rounding is half-up rather than Python's default banker's rounding --
    round(32.5) returns 32 in Python, which silently lands a player in the
    wrong third of a round. At a boundary this is a coin flip anyway:
    being one pick off is noise next to a typical adp_stdev of 4-6 picks.
    """
    pick = max(1, int(math.floor(overall_pick + 0.5)))
    round_number = math.ceil(pick / teams)
    pick_in_round = pick - (round_number - 1) * teams

    if pick_in_round <= teams / 3:
        third = "Beginning of"
    elif pick_in_round <= 2 * teams / 3:
        third = "Middle of"
    else:
        third = "End of"
    return f"{third} Round {round_number}"


def format_pick(overall_pick, teams):
    """
    Turns an overall pick number into compact 'Rd.Pk' form: 13th overall is
    '3.01' in a 6-team league and '2.01' in a 12-team one.

    This is why the board stores ADP as an overall pick number and formats it
    at write time rather than shipping FFC's own `adp_formatted` string. FFC
    returns Rd.Pk already rendered for a 12-team draft, which is silently
    wrong on a 6-team board -- '2.01' would print next to a player who
    actually goes in the third round of your draft.

    The ORDERING is league-agnostic and transfers fine; the round labels are
    not. See build_notes() for the caveat that survives this conversion.
    """
    pick = max(1, int(math.floor(overall_pick + 0.5)))
    round_number = math.ceil(pick / teams)
    pick_in_round = pick - (round_number - 1) * teams
    return f"{round_number}.{pick_in_round:02d}"


def compute_draft_targets(players, teams):
    """
    Adds `rank`, `value_delta`, and `draft_target`.

    Sort order is (has_adp, vor) descending. Players with no real ADP are
    hard-capped below every ADP-bearing player regardless of how good the
    stats say they are -- if the market has no opinion on someone, the
    model's opinion alone isn't worth a pick you could spend on a known
    quantity.

    Draft target answers "when do I actually take him," which is not the
    same as "how good is he":

      - Bargain (value_delta >= 0, model likes him more than the market):
        target = his real ADP minus one standard deviation of that ADP.
        You wait as long as you safely can instead of reaching to his
        pure-value rank, and the cushion is his OWN volatility -- a
        consensus player needs less room than a divisive one.

      - Market premium (value_delta < 0): show 'Fair value: Round X' from
        the model's own rank. Don't pay up for what the stats don't
        support; take him only if he actually falls that far.

      - No ADP: plain model-rank slot, since there's no market to compare.
    """
    # Ties on VOR break by ADP (earlier ADP wins), then by player_name.
    #
    # ADP as a TIEBREAKER does not violate the statistics-only rule --
    # same standing as depth charts in FEATURE_SPEC.md. It never enters
    # adjusted_fantasy_points_per_game or vor, so it cannot move a player
    # past anyone the model actually separated. It only decides the order
    # of players the model rated *identically*, where the alternative is
    # alphabetical, i.e. arbitrary.
    #
    # Exact VOR ties are real right now: rookies sharing a position/round
    # cohort get byte-identical projections (Jeremiyah Love and Jadarian
    # Price), so some third key is required for a reproducible board.
    # player_name stays as a final key because no-ADP players have a null
    # adp and would otherwise still be nondeterministic among themselves.
    # Phase 12 removes the ties themselves.
    # out_for_season sorts FIRST, so those players land below everyone --
    # below even the no-ADP block. They keep a real VOR so you can see
    # what they'd have been worth, but they can never surface as a pick.
    players = players.sort(
        ["out_for_season", "has_adp", "vor", "adp", "player_name"],
        descending=[False, True, True, False, False],
        nulls_last=True,
    ).with_row_index("rank", offset=1)

    # value_delta = where the market takes him minus where the model ranks
    # him. Positive = the model is higher on him than the room is.
    players = players.with_columns(
        pl.col("adp").rank(method="dense").cast(pl.Int32).alias("adp_rank")
    ).with_columns(
        (pl.col("adp_rank") - pl.col("rank").cast(pl.Int32)).alias("value_delta")
    )

    rows = players.to_dicts()
    targets = []
    adp_slots = []
    for row in rows:
        # ADP re-expressed in THIS league's draft shape, not FFC's 12-team one.
        if row["has_adp"] and row["adp"] is not None:
            adp_slots.append(format_pick(row["adp"], teams))
        else:
            adp_slots.append(None)

        if row["out_for_season"]:
            # No round, ever. The whole point is that this player must not
            # read as takeable at any price -- and his value_delta gets
            # blanked in write_workbook for the same reason.
            targets.append("DO NOT DRAFT — out for season")
        elif not row["has_adp"]:
            targets.append(format_slot(row["rank"], teams))
        elif row["value_delta"] is not None and row["value_delta"] >= 0:
            cushion = row["adp_stdev"] or 0.0
            targets.append(format_slot(row["adp"] - cushion, teams))
        else:
            targets.append("Fair value: " + format_slot(row["rank"], teams))

    return players.with_columns([
        pl.Series("draft_target", targets),
        pl.Series("adp_slot", adp_slots, dtype=pl.String),
    ])


def build_notes(replacement_ranks, teams, rounds, config=None, starter_ranks=None):
    """The explanatory block at the top of the sheet. It lives here so the
    board always ships with an accurate description of how it was built."""
    order = [p for p in MODELED_POSITIONS if p in replacement_ranks]
    levels = " / ".join(f"{p}{replacement_ranks[p]}" for p in order)

    season = (config or {}).get("fantasy_season_length")
    weeks = (config or {}).get("regular_season_weeks")

    unmodeled = unmodeled_slots_per_team(config or {"roster_slots": {}})

    replacement_note = ""
    if starter_ranks:
        old = " / ".join(f"{p}{starter_ranks[p]}" for p in order
                         if p in starter_ranks)
        replacement_note = (
            f"CHANGED IN v11: replacement level used to be the last STARTER ({old}) and is now "
            f"the last player DRAFTED ({levels}) -- {teams * rounds} picks minus "
            f"{teams * unmodeled} for K/DST, split by the position mix of that "
            "many players in ADP order. The old rule was defensible at 12 teams and wrong at 6: "
            "it set replacement at QB6 in a league where QB7 through QB32 are all on waivers, so "
            "the board was telling you to spend an early pick on a quarterback you could stream. "
            "Every VOR on this sheet is larger than it was in v10 because the bar moved down; "
            "what matters is that it moved down FURTHER at RB and WR than at QB and TE.\n"
        )

    availability_note = ""
    if season:
        window = f"weeks {weeks[0]}-{weeks[1]}" if weeks else f"{season:g} weeks"
        availability_note = (
            f"Exp Gm / Exp Pts = availability, deliberately kept OUT of Rank and VOR. Exp Gm is "
            f"this league's REGULAR season ({window}, {season:g} games) minus known PUP/NFI "
            "absence; Exp Pts is Adj PPG x Exp Gm. The denominator is the regular season, not "
            "the 17-week NFL calendar, because games after your league's final are worth nothing. "
            "A four-game PUP absence therefore costs a different share here than in a league with "
            "a different schedule, which is why the same player can be worth more in one of your "
            "two leagues than the other. Adj PPG stays a pure per-game rate: an injury makes a "
            "player play fewer games, not play worse in the ones he plays.\n"
        )

    gaps, slot = league_gaps(config or {"num_teams": teams})
    slot_note = ("draft slot %d" % slot if (config or {}).get("draft_slot")
                 else "draft slot unknown, assuming %d (middle)" % slot)
    vona_note = (
        f"HOW TO USE THIS SHEET (Phase 14). Do not draft straight down Rank. A simulated "
        f"bakeoff over 120 drafts per league found pure best-available finished LAST of the "
        f"five model-based strategies in all four leagues, because VOR is computed once "
        f"against an EMPTY roster and never learns what you already hold -- so it keeps "
        f"pricing your fifth running back as if he were your first.\n"
        f"Instead: among positions where a STARTING slot is still empty, take the player with "
        f"the biggest gap between his Adj PPG and the best man at his position you expect to "
        f"survive to your next pick. Cap QB at 2 and TE at 2 (QB 3 in superflex). If every gap "
        f"is near zero, nothing is scarce -- then, and only then, take the best Rank.\n"
        f"THE CALCULATION, done live. Your two snake gaps in this league are {gaps[0]} and "
        f"{gaps[-1]} picks ({slot_note}) -- long gap in odd rounds, short gap in even ones. On the "
        f"clock: for each position where a starting slot is still empty, take the best man left "
        f"(sort by PPG@Pos), then count how many at that position you expect to go before your "
        f"next pick and look that many rows further down. The difference in Adj PPG is what "
        f"passing costs you. Take the biggest one.\n"
        f"COUNT FROM WHAT YOU HAVE WATCHED, not from ADP (Ovr). ADP is the starting estimate and "
        f"the room leaves it immediately; if four backs went in the last six picks, count four. "
        f"That correction is the whole game. A precomputed version of this column shipped briefly "
        f"in v17.1 and was cut: drafting off the frozen numbers scored +18.8 points against "
        f"best-available where the same rule computing it live scored +59.7. Two thirds of the "
        f"edge is in the correction a frozen column cannot make.\n"
        f"ADP (Ovr) is the right column for \"will he last until my next pick\" -- it is a raw "
        f"pick number, so compare it to your next pick directly.\n"
        "Draft Target and Value Δ moved to the CHECK block and are REFERENCE, not instruction. "
        "Do NOT use Draft Target to judge whether a player lasts to your next pick: it is ADP "
        "minus one standard deviation rounded into thirds of a round, and for players the market "
        "likes more than the model it switches meaning entirely to \"Fair value: Round X\" off "
        "model rank. Two quantities in one column, both bucketed. ADP (Ovr) answers that question "
        "in raw picks.\n"
    )
    return (
        vona_note
        + "Rank/VOR = who's best by the stats-only model. VOR = points per game above that "
        f"position's replacement level ({levels}), which is why a 26-PPG quarterback does not "
        "outrank a 20-PPG receiver -- in a 1-QB league the gap over the freely available "
        "alternative is what a pick actually buys.\n"
        + replacement_note
        + availability_note
        + "Why (value drivers) = what moved him off an AVERAGE player at his position, biggest "
        "effect first, in points per game. These are the actual terms of the fitted model, not a "
        "description of it: they sum to the Sit Adj column beside them. '62% team share' reading "
        "negative is not a typo -- the model predicts the CHANGE from a player's own trailing "
        "baseline, and a back already holding most of his backfield has that usage priced in.\n"
        "Draft Target = when to actually take him. For bargains (Value Δ >= 0) it is real ADP "
        "minus a one-standard-deviation cushion (his own adp_stdev), so you wait as late as you "
        "safely can rather than reaching to pure value. For players the market likes MORE than "
        "the model (Value Δ < 0) it shows \"Fair value: Round X\" -- the model's own rank. "
        "Don't pay a premium the stats don't support; take him only if he falls that far.\n"
        "Value Δ = ADP rank minus model rank; positive = model likes him more than the market. "
        "Sit Adj = the situational adjustment applied to his raw statistical baseline; it is now "
        "two-sided (v7 and earlier were negative for every player -- the regression intercept was "
        "missing from the code, forcing a uniform penalty of 3.59 RB / 2.37 WR / 1.58 TE).\n"
        f"Gray rows = ranked past pick {teams * rounds}, the last pick of a {teams}-team "
        f"{rounds}-round draft. Pink rows = no real ADP, hard-capped below every ADP-bearing "
        "player regardless of stats. Recent Injury (red) = ended the 2025 regular season on IR, "
        "reference only.\n"
        "DARK RED rows = flagged OUT_SEASON in injury_overrides.csv: sorted below everything, "
        "Value Δ blanked, no draft round shown. That file is maintained by hand because the "
        "model cannot see 2026 preseason injuries -- and because refreshing ADP without it is "
        "actively harmful: the market drops an injured player instantly while the model's "
        "trailing-average projection does not move, so the gap gets labelled a bargain.\n"
        "Rookies (Rook = R) take no situational adjustment and share one cohort baseline per "
        "position/round, so two rookies of the same draft round can tie exactly. K and DST are "
        "not modeled; draft those separately.\n"
        + adp_caveat(teams, mock_sourced=bool((config or {}).get("adp_mock_file")))
    )


def adp_caveat(teams, source_teams=12, mock_sourced=False):
    """
    The honest label on the ADP columns.

    ADP (Ovr) is an overall pick number and ADP (Rd.Pk) re-expresses it in this
    league's draft shape. Both are derived from FFC mock drafts run at
    `source_teams`. Rescaling preserves the ORDER players come off the board;
    it does not reproduce the BEHAVIOR of a different-sized draft. In a 6-team
    league only half as many players are taken, so scarcity at QB and TE
    largely disappears and those positions genuinely go later than a rescaled
    12-team ADP implies. The column is a reference for order, not a prediction
    of when someone goes in your draft.
    """
    if mock_sourced:
        # Phase 13.6. No rescaling happened and none is needed: these picks
        # were made in this league's own draft shape, so the caveat the rest
        # of this function exists to give does not apply. The one that
        # replaces it is about sample size, which is the honest weakness of
        # two drafts.
        return (
            f"ADP (Ovr) = average overall pick across two real {teams}-team "
            f"superflex mock drafts (no rescaling -- these picks were made in "
            f"this draft's own shape, unlike the FFC feed every other board "
            f"here reads). ADP (Rd.Pk) is that number as {format_pick(13, teams)}-style "
            f"Rd.Pk. CAUTION: two drafts is a thin sample next to FFC's "
            f"hundreds, and one of the two had most teams autodrafting. A "
            f"player taken in only ONE of the two is averaged against pick "
            f"{teams * 12 + 1} -- going undrafted in a {teams}-team room is "
            f"evidence, not a missing value -- so his ADP sits deliberately "
            f"deeper than his one observed pick and his Draft Target cushion "
            f"widens to match."
        )

    base = (
        f"ADP (Ovr) = overall pick number in {source_teams}-team FFC mock drafts. "
        f"ADP (Rd.Pk) is that same number re-expressed for a {teams}-team draft "
        f"(pick 13 reads {format_pick(13, teams)} here)."
    )
    if teams == source_teams:
        return base
    return (
        base + " CAUTION: rescaling preserves the ORDER players go, not the "
        f"BEHAVIOR of a {teams}-team draft. Only {teams} of every {source_teams} "
        "picks happen here, so positional scarcity is much weaker -- QB and TE in "
        "particular should be expected to last longer than this column suggests. "
        "Treat it as a consensus ranking, not a predicted draft slot."
    )


def write_workbook(board, replacement_ranks, config, output_path, build_note=None,
                   starter_ranks=None):
    """Writes the formatted sheet. Values only -- no formulas, since every
    number here is a model output that would be wrong if a user edited a
    cell and Excel recalculated something downstream from it."""
    teams = config["num_teams"]
    rounds = config["total_rounds"]
    last_pick = teams * rounds

    wb = Workbook()
    ws = wb.active
    ws.title = "Draft Board"

    n_cols = len(COLUMNS)
    last_col = get_column_letter(n_cols)

    # Title
    ws.merge_cells(f"A1:{last_col}1")
    title = ws["A1"]
    title.value = (
        f"{config['league_name']} — 2026 Draft Board "
        f"({teams}-team, {rounds} rounds, statistics-only model, VOR draft order)"
    )
    title.font = Font(name=FONT_NAME, size=14, bold=True)
    ws.row_dimensions[1].height = 22

    # Notes block
    # Phase 14 grew this block by roughly a screen of text -- the whole
    # "how to use this sheet" rule now lives here, because a board that
    # ships without it invites exactly the best-available drafting the
    # simulator just measured as the worst strategy available. Excel does
    # not auto-fit merged cells, so the row span and the row heights are
    # both set explicitly; getting one without the other silently clips
    # the instructions.
    ws.merge_cells(f"A2:{last_col}{NOTES_LAST_ROW}")
    for _r in range(2, NOTES_LAST_ROW + 1):
        ws.row_dimensions[_r].height = 58
    notes = ws["A2"]
    notes.value = build_notes(replacement_ranks, teams, rounds, config, starter_ranks)
    notes.font = Font(name=FONT_NAME, size=9, color="555555")
    notes.alignment = Alignment(wrap_text=True, vertical="top")

    # Header
    header_row = NOTES_LAST_ROW + 2
    thin_bottom = Border(bottom=Side(style="thin"))
    for i, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(header_row, i, label)
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_bottom
        ws.column_dimensions[get_column_letter(i)].width = width

    # Data
    for offset, row in enumerate(board.to_dicts()):
        r = header_row + 1 + offset
        rank = row["rank"]
        has_adp = bool(row["has_adp"])

        out_for_season = bool(row["out_for_season"])

        if out_for_season:
            fill_color = OUT_FILL
        elif not has_adp:
            fill_color = NO_ADP_FILL
        elif rank > last_pick:
            fill_color = UNDRAFTABLE_FILL
        else:
            fill_color = POSITION_FILLS.get(row["position"], NO_ADP_FILL)
        fill = PatternFill("solid", start_color=fill_color)

        # Value Δ is blanked for out players. It would otherwise read
        # hugely positive -- the market drops an injured player while the
        # model's trailing-average projection doesn't move -- and that gap
        # is exactly what the column normally labels "bargain."
        # Keyed by column LABEL, not position. Reordering COLUMNS above is
        # now a one-line edit -- previously it meant hand-shuffling this
        # list and a block of number-format indices in lockstep, with
        # nothing failing loudly if you got it wrong.
        cells = {
            "Rank": rank,
            "Pos": row["position"],
            "Player": row["player_name"],
            "Draft Target": row["draft_target"],
            "VOR": row["vor"],
            "Adj PPG": row["adjusted_fantasy_points_per_game"],
            "PPG@Pos": row.get("ppg_pos_rank"),
            "Value Δ (picks)": (row["value_delta"]
                                if (has_adp and not out_for_season) else None),
            "ADP (Rd.Pk)": row["adp_slot"] if has_adp else None,
            "ADP (Ovr)": row["adp"] if has_adp else None,
            "Why (value drivers)": row.get("value_drivers"),
            "Bye": row["bye"] if has_adp else None,
            "Team": row["team"],
            "Recent Injury": ("OUT (2026)" if out_for_season
                              else ("INJURED" if row["recent_major_injury"] else None)),
            "Exp Gm": row.get("expected_games"),
            "Exp Pts": row.get("expected_total_points"),
            "Sit Adj": row.get("situational_adjustment"),
            "Rook": "R" if row["is_rookie"] else None,
            "Age": row.get("age"),
            # Stored as a share-per-season slope (0.021 = +2.1 points of
            # team share per year); rendered in percentage points, which
            # is the unit the number is actually legible in.
            "Usage Trend": (row.get("usage_trend_share") * 100.0
                            if row.get("usage_trend_share") is not None else None),
            "Trend n": row.get("trend_seasons_used"),
            "GP (sample)": row["games_played"],
            "Has ADP": has_adp,
            # Override note, else blank for draft-day scribbling.
            "Notes (manual)": row.get("injury_note"),
        }
        missing = [label for label, _ in COLUMNS if label not in cells]
        if missing:
            raise KeyError(f"COLUMNS declares {missing} with no value supplied")
        values = [cells[label] for label, _ in COLUMNS]

        # Left-align the text columns; everything else centers.
        left_aligned = {
            COLUMN_INDEX["Player"],
            COLUMN_INDEX["Why (value drivers)"],
            COLUMN_INDEX["Notes (manual)"],
        }
        for i, value in enumerate(values, start=1):
            cell = ws.cell(r, i, value)
            cell.fill = fill
            cell.font = Font(name=FONT_NAME, size=11)
            cell.alignment = Alignment(
                horizontal="left" if i in left_aligned else "center"
            )

        # Looked up by NAME, not position. Phase 11 adds three columns and
        # the old hardcoded indices would all have shifted silently -- the
        # number format is cosmetic, so nothing would have failed loudly.
        for label, fmt in (
            ("Adj PPG", "0.0"),
            ("Sit Adj", "+0.0;-0.0;0.0"),
            ("VOR", "0.0"),
            ("Exp Gm", "0.0"),
            ("Exp Pts", "0"),
            ("ADP (Ovr)", "0.0"),
            ("Value Δ (picks)", "+0;-0;0"),
            ("Age", "0.0"),
            ("Usage Trend", "+0.0;-0.0;0.0"),
        ):
            ws.cell(r, COLUMN_INDEX[label]).number_format = fmt

        # "Has ADP" is kept for filtering but rendered invisible -- white
        # text on the row fill -- so it doesn't add visual noise.
        ws.cell(r, COLUMN_INDEX["Has ADP"]).font = Font(
            name=FONT_NAME, size=11, color="FFFFFF"
        )

        if out_for_season:
            for label in ("Player", "Draft Target", "Recent Injury"):
                cell = ws.cell(r, COLUMN_INDEX[label])
                cell.fill = PatternFill("solid", start_color=INJURY_FILL)
                cell.font = Font(name=FONT_NAME, size=11, color="FFFFFF", bold=True)
        elif row["recent_major_injury"]:
            injury = ws.cell(r, COLUMN_INDEX["Recent Injury"])
            injury.fill = PatternFill("solid", start_color=INJURY_FILL)
            injury.font = Font(name=FONT_NAME, size=11, color="FFFFFF", bold=True)

        # Phase 11 B (CP4). Rookies are skipped: their baseline is a
        # cohort projection and `games_played` describes a college career
        # the model never saw, so shading it would be answering a
        # different question than the one the colour implies.
        games = row.get("games_played")
        if games is not None and not row["is_rookie"]:
            if games < VERY_LOW_CONFIDENCE_GAMES:
                confidence_fill = VERY_LOW_CONFIDENCE_FILL
            elif games < LOW_CONFIDENCE_GAMES:
                confidence_fill = LOW_CONFIDENCE_FILL
            else:
                confidence_fill = None
            if confidence_fill:
                cell = ws.cell(r, COLUMN_INDEX["GP (sample)"])
                cell.fill = PatternFill("solid", start_color=confidence_fill)
                cell.font = Font(name=FONT_NAME, size=11, bold=True)

    last_row = header_row + board.height
    ws.freeze_panes = f"D{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"

    output_path = Path(output_path)
    write_build_history(wb, output_path, config, build_note)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
    except PermissionError:
        # Excel holds an exclusive lock on an open workbook, so a rebuild
        # over a board you are looking at dies with a bare traceback from
        # deep inside zipfile. Harmless, and the least helpful possible
        # moment for it: the one time this is guaranteed to happen is
        # draft day, refreshing ADP with the board open on the other
        # monitor and a pick clock running.
        #
        # Everything upstream has already succeeded by this point -- the
        # model ran, the gate passed, replacement levels resolved. Only
        # the file write failed, and re-running costs nothing.
        raise SystemExit(
            f"\nCannot write {output_path.name} -- it is open in Excel "
            f"(or another program has it locked).\n"
            f"Close it and re-run. Nothing else went wrong: the board built "
            f"fine and only the save failed.\n"
        ) from None
    return output_path


def write_build_history(workbook, output_path, config, build_note):
    """
    Appends this build to the in-workbook log, carrying prior rows forward.

    Read the prior history BEFORE the new file is saved over it -- the read
    happens here rather than in build_board() so the two can't drift apart.
    """
    prior = read_build_history(output_path)
    sheet = workbook.create_sheet(HISTORY_SHEET)

    headers = ["Build", "Built (local time)", "Model ver", "Git", "League", "Note"]
    for i, label in enumerate(headers, start=1):
        cell = sheet.cell(1, i, label)
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=HEADER_FILL)
        sheet.column_dimensions[get_column_letter(i)].width = (
            8 if i == 1 else 20 if i in (2, 5) else 11 if i in (3, 4) else 60
        )

    for offset, row in enumerate(prior):
        for i, value in enumerate(row[: len(headers)], start=1):
            sheet.cell(2 + offset, i, value).font = Font(name=FONT_NAME, size=10)

    new_row = 2 + len(prior)
    values = [
        len(prior) + 1,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        MODEL_VERSION,
        git_short_hash(),
        config["league_name"],
        build_note or "",
    ]
    for i, value in enumerate(values, start=1):
        cell = sheet.cell(new_row, i, value)
        cell.font = Font(name=FONT_NAME, size=10, bold=True)
    sheet.cell(new_row, 6).alignment = Alignment(wrap_text=True, vertical="top")


def prepare_board_frame(features_path, config, quiet=False):
    """
    Everything between player_features.csv and a ranked board, with no
    spreadsheet involved.

    EXTRACTED FROM build_board() (Aug 6), and the reason is the same one
    that came up three times today: anything that wants to inspect the
    board has to get it from the code that builds the board, not from a
    second implementation that agrees until it doesn't. `sanity_top_n`
    calls this; so does build_board. There is one order of operations and
    both see it.

    Returns (board, replacement_ranks, starter_ranks).
    """
    players = pl.read_csv(features_path).filter(
        pl.col("position").is_in(MODELED_POSITIONS)
    )

    players = select_adp_variant(players, config)

    # BEFORE anything reads a PPG column. Replacement level, VOR, draft
    # targets and the sheet itself all descend from these numbers, so a
    # rescore applied later would leave some of the board in one scoring
    # system and some in another -- which is worse than the bug it fixes.
    players = rescore_for_league(players, config)

    # CSV round-trips booleans as "true"/"false" strings; normalize once
    # here so every downstream check is a real boolean.
    for column in ["has_adp", "is_rookie", "recent_major_injury"]:
        players = players.with_columns(
            pl.col(column).cast(pl.String).str.to_lowercase().eq("true").alias(column)
        )

    # AFTER the boolean cast, because this rewrites `has_adp` itself, and
    # BEFORE anything reads ADP -- replacement level, Value delta, draft
    # targets and the row sort all descend from it.
    players = apply_mock_adp(players, config)

    players = apply_injury_overrides(players)

    # PHASE 13.5. Rookie availability, and it lands HERE on purpose --
    # after the injury overrides have populated `expected_games_missed`,
    # before `compute_expected_points` consumes it. One column, one
    # consumer, and a rookie who is also on PUP takes the larger of the
    # two absences rather than their sum.
    #
    # WHAT THIS DOES NOT DO. The Phase 13.5 gate compared rate@8 and
    # rate@0 against an availability term and chose availability ALONE:
    # rate@0 added 0.05 SE of RMSE while turning an unbiased predictor
    # (+1.30 pts) into a biased one (-6.29). Lowering the games threshold
    # is an availability correction wearing a rate's clothing, and
    # applying it alongside an explicit availability term double-counts --
    # exactly as compute_expected_points warned it would. So no cohort
    # baseline changes, `adjusted_fantasy_points_per_game` is untouched,
    # and NO RANK MOVES. This affects Exp Pts and nothing else.
    playing_time = load_playing_time_model()
    if playing_time is None:
        if not quiet:
            print("No data/playing_time.json -- rookies keep a full season of "
                  "expected games. Run `python -m src.playing_time`.")
    else:
        players = expected_games_for_rookies(players, playing_time, config)
        if not quiet:
            # EXCLUDES out-for-season rookies, and the exclusion is the
            # point. `compute_expected_points` zeroes an OUT_SEASON player
            # regardless of what availability says, so counting him here
            # would print 232 against a board where 231 rows actually
            # moved -- and this number exists precisely to be checked
            # against `compare_boards`. A count that cannot be reconciled
            # is worse than no count.
            #
            # It was 232 on the first v14 build. The extra man was Chris
            # Brazzell II, a rookie WR (rd 3, pick 83) marked OUT_SEASON,
            # whose Exp Gm was already 0 and stayed 0.
            adjusted = players.filter(
                pl.col("is_rookie")
                & (pl.col("expected_games_missed") > 0)
                & ~pl.col("out_for_season")
            )
            print(f"Rookie availability: {adjusted.height} rookies marked down "
                  f"on Exp Pts (rank unaffected by construction).")

    players = compute_expected_points(players, config)

    # How far shrinkage moved this player, for the Why column. Computed
    # here rather than stored in player_features.csv so it can never
    # disagree with the two columns it is the difference of.
    if "fantasy_points_per_game_shrunk" in players.columns:
        players = players.with_columns(
            (pl.col("fantasy_points_per_game_shrunk")
             - pl.col("fantasy_points_per_game")).alias("baseline_shrink_delta")
        )
        moved = players.filter(pl.col("baseline_shrink_delta").abs() >= 0.5)
        if not quiet:
            print(f"Baseline shrinkage: {moved.height} players moved by 0.5+ PPG")
    elif not quiet:
        print("NOTE: no shrunk baseline in player_features.csv -- "
              "re-run `python -m src.pipeline` to apply Phase 11 CP5.")

    players = build_value_drivers(players)

    starter_ranks = compute_starter_ranks(config)
    replacement_ranks = compute_replacement_ranks(config, players)
    if not quiet:
        print(f"Replacement ranks: {replacement_ranks} "
              f"(was, by starter slots: {starter_ranks})")

    board = compute_vor(players, replacement_ranks)
    board = compute_draft_targets(board, config["num_teams"])
    board = add_vona_columns(board, config)
    return board, replacement_ranks, starter_ranks


def build_board(features_path=FEATURES_PATH, output_path=None,
                config_path=CONFIG_PATH, version=MODEL_VERSION, build_note=None,
                skip_gate=False):
    require_holdout_gate(skip=skip_gate)

    config = load_config(config_path)
    teams = config["num_teams"]

    board, replacement_ranks, starter_ranks = prepare_board_frame(
        features_path, config
    )

    if output_path is None:
        output_path = PROJECT_ROOT / f"2026_{league_slug(config)}_Board_v{version}.xlsx"

    written = write_workbook(board, replacement_ranks, config, output_path, build_note,
                             starter_ranks=starter_ranks)
    history_count = len(read_build_history(written))
    print(f"Wrote {board.height} players to {written}")
    print(f"{config['league_name']} — {teams} teams, model v{version}, "
          f"build #{history_count} ({git_short_hash()})")

    top = board.head(10).select(
        ["rank", "position", "player_name", "adjusted_fantasy_points_per_game",
         "vor", "draft_target"]
    )
    print("\nTop 10:")
    print(top)
    return board


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build the 2026 Excel draft board.")
    parser.add_argument("--config", type=str, default=str(CONFIG_PATH),
                        help="league config json (default: league_config_12team.json, "
                             "i.e. the 12-team league). Use league_config_6team.json "
                             "for the 6-team board.")
    parser.add_argument("--version", type=int, default=MODEL_VERSION,
                        help="model version for the output filename; bump only "
                             "when the MODEL changes, not for a data refresh")
    parser.add_argument("--output", type=str, default=None,
                        help="explicit output path (overrides the derived name)")
    parser.add_argument("--features", type=str, default=str(FEATURES_PATH),
                        help="path to player_features.csv")
    parser.add_argument("--note", type=str, default=None,
                        help="what changed in this build; recorded in the "
                             "Build History sheet")
    parser.add_argument("--skip-gate", action="store_true",
                        help="build even if the holdout gate is missing, stale "
                             "or failing. For emergencies only -- the resulting "
                             "board may rank players on features that are known "
                             "not to predict out of sample.")
    args = parser.parse_args()

    build_board(
        features_path=args.features,
        output_path=args.output,
        config_path=args.config,
        version=args.version,
        build_note=args.note,
        skip_gate=args.skip_gate,
    )
