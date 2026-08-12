"""
Diffs two draft boards, and asserts the thing a build was supposed to do.

WHY THIS EXISTS
---------------
Phase 13.5 ships a change with a precise, checkable claim attached: it
touches `expected_games` for rookies and therefore Exp Pts, and it moves
NO player's rank anywhere. That claim is either true of the built file or
it is not, and "I looked at the top 30 and it seemed fine" is not how you
find out -- a rank change at row 400 is exactly the kind of thing a human
read misses, and it is exactly the kind of thing a mis-wired hook causes.

So the claim gets executed instead of asserted. `--expect rank-identical`
exits non-zero if a single rank moved.

It generalises past this one phase. Every future build is a claim about
what should and should not have changed, and the two useful modes are
already here:

    --expect rank-identical   a change that must not touch rank
                              (Exp Pts work, cosmetic columns, a
                              scoring-neutral refactor)
    --expect any              a data refresh, where ranks SHOULD move and
                              the question is only whether they moved
                              sanely and where

WHAT IT DELIBERATELY DOES NOT DO. It does not judge whether a new ranking
is better. `holdout.py` and `playing_time.py --gate` do that, out of
sample, before a board is allowed to exist. This runs after, and it only
answers "did the file change in the way the change was supposed to."

USAGE
-----
    python -m src.compare_boards OLD.xlsx NEW.xlsx
    python -m src.compare_boards OLD.xlsx NEW.xlsx --expect rank-identical
    python -m src.compare_boards OLD.xlsx NEW.xlsx --top 60
"""

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

# The board writes a title block and a legend above the table, so the
# header row is found rather than assumed -- a layout change upstream
# should not silently shift every column by one.
HEADER_ANCHOR = "Rank"
MAX_HEADER_SCAN = 40

# Columns whose movement is the POINT of a Phase 13.5-style build, and
# columns whose movement would mean something went wrong. Reported
# separately so the two never blur together in the output.
EXPECTED_TO_MOVE = ["Exp Gm", "Exp Pts"]
MUST_NOT_MOVE = ["Rank", "VOR", "Adj PPG", "Pos"]


def read_board(path):
    """One dict per player row, keyed by the header labels."""
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["Draft Board"]

    header_row = None
    for row in range(1, MAX_HEADER_SCAN):
        if sheet.cell(row, 1).value == HEADER_ANCHOR:
            header_row = row
            break
    if header_row is None:
        raise SystemExit(
            f"{Path(path).name}: no '{HEADER_ANCHOR}' header in the first "
            f"{MAX_HEADER_SCAN} rows. Is this a draft board?"
        )

    headers = [
        sheet.cell(header_row, column).value
        for column in range(1, sheet.max_column + 1)
    ]

    rows = {}
    for row in range(header_row + 1, sheet.max_row + 1):
        values = {
            header: sheet.cell(row, column).value
            for column, header in enumerate(headers, start=1)
            if header
        }
        if values.get("Rank") is None or not values.get("Player"):
            continue
        rows[values["Player"]] = values
    return rows


def _delta(old, new):
    """Numeric difference, or None when either side is not a number."""
    if isinstance(old, (int, float)) and isinstance(new, (int, float)):
        return new - old
    return None


def compare(old_path, new_path, top=60, expect="any", focus=200):
    old = read_board(old_path)
    new = read_board(new_path)

    print(f"OLD  {Path(old_path).name}  {len(old)} players")
    print(f"NEW  {Path(new_path).name}  {len(new)} players")

    added = sorted(set(new) - set(old))
    dropped = sorted(set(old) - set(new))
    if added:
        print(f"\nADDED ({len(added)}): {', '.join(added[:12])}"
              f"{' ...' if len(added) > 12 else ''}")
    if dropped:
        print(f"DROPPED ({len(dropped)}): {', '.join(dropped[:12])}"
              f"{' ...' if len(dropped) > 12 else ''}")

    shared = [player for player in new if player in old]

    moved = []
    for player in shared:
        change = _delta(old[player].get("Rank"), new[player].get("Rank"))
        if change:
            moved.append((abs(change), change, player))
    moved.sort(reverse=True)

    # THE RANK REPORT IS USELESS WITHOUT THESE TWO THINGS, and the Aug 12
    # refresh is why both were added.
    #
    # 1. A DEPTH FILTER. That run reported "878 of 1082 moved" and led with
    #    Owen Wright +481 -- rank 589 to 1070, a player nobody will draft in
    #    any of the three leagues. Meanwhile the fact that exactly ONE
    #    player inside the top 120 moved by 10 or more was nowhere on the
    #    screen. Deep churn is loud and meaningless; `--focus` puts the
    #    draftable range first and relegates the rest to a count.
    #
    # 2. AN ADP-FLIP MARKER. `compute_draft_targets` sorts by
    #    (out_for_season, has_adp, vor, ...), so `has_adp` is a hard gate
    #    ABOVE vor -- gaining ADP coverage vaults a player over the entire
    #    no-ADP block no matter what the model thinks. On that same run all
    #    FIVE of the biggest top-200 moves were coverage flips with a PPG
    #    change of essentially zero: Stefon Diggs 188 -> 91 on dPPG -0.24,
    #    and three others on dPPG of exactly 0.00.
    #
    #    Without the marker those read as the model radically revaluing
    #    four players. They are the FFC feed adding and dropping names. The
    #    distinction decides whether you act on the move.
    focused = [entry for entry in moved
               if min(old[entry[2]]["Rank"], new[entry[2]]["Rank"]) <= focus]
    deep = len(moved) - len(focused)

    def adp_flag(player):
        was, now = old[player].get("Has ADP"), new[player].get("Has ADP")
        if was == now:
            return ""
        return "  GAINED ADP" if now else "  LOST ADP"

    print(f"\nRANK: {len(moved)} of {len(shared)} shared players moved "
          f"({len(focused)} inside the top {focus}, {deep} deeper).")
    if focused:
        print(f"  {'player':24s} {'old':>5s} {'new':>5s} {'move':>6s} {'dPPG':>7s}")
        for _, change, player in focused[:20]:
            ppg = _delta(old[player].get("Adj PPG"), new[player].get("Adj PPG"))
            print(f"  {player[:24]:24s} {old[player]['Rank']:>5} "
                  f"{new[player]['Rank']:>5} {change:>+6.0f} "
                  f"{ppg if ppg is not None else 0.0:>+7.2f}{adp_flag(player)}")
        if len(focused) > 20:
            print(f"  ... and {len(focused) - 20} more inside the top {focus}")
    print("  A large move with dPPG near zero is a sort-order effect, not a "
          "revaluation -- check the ADP flag.")

    for column in EXPECTED_TO_MOVE + MUST_NOT_MOVE:
        if column == "Rank":
            continue
        changed = []
        for player in shared:
            change = _delta(old[player].get(column), new[player].get(column))
            if change is not None and abs(change) > 1e-9:
                changed.append((abs(change), change, player))
        if not changed:
            print(f"\n{column}: unchanged for all {len(shared)} players.")
            continue
        changed.sort(reverse=True)
        # The board writes "R" in the Rook column, not "Y". The first
        # version of this check looked for "Y" and reported "0 of them
        # rookies" against a build that had just printed "232 rookies
        # marked down" two lines earlier. Any non-empty marker counts now,
        # so the column can change its mind without this lying about it.
        rookies = sum(
            1 for _, _, player in changed
            if str(new[player].get("Rook") or "").strip()
        )
        print(f"\n{column}: {len(changed)} changed "
              f"({rookies} of them rookies). Largest:")
        for _, change, player in changed[:8]:
            print(f"  {player[:24]:24s} {old[player].get(column)} -> "
                  f"{new[player].get(column)}   ({change:+.2f})")

    print(f"\nTOP {top} POSITION MIX")
    for label, board in (("old", old), ("new", new)):
        counts = {}
        for values in board.values():
            if isinstance(values.get("Rank"), (int, float)) and values["Rank"] <= top:
                counts[values.get("Pos")] = counts.get(values.get("Pos"), 0) + 1
        print(f"  {label}: " + "  ".join(
            f"{position} {counts.get(position, 0)}"
            for position in ["QB", "RB", "WR", "TE"]
        ))

    if expect == "rank-identical":
        problems = []
        if moved:
            problems.append(f"{len(moved)} ranks moved")
        if added or dropped:
            problems.append(f"{len(added)} added, {len(dropped)} dropped")
        for column in MUST_NOT_MOVE:
            if column == "Rank":
                continue
            for player in shared:
                change = _delta(old[player].get(column), new[player].get(column))
                if change is not None and abs(change) > 1e-9:
                    problems.append(f"{column} moved (e.g. {player})")
                    break
        if problems:
            print("\nFAILED --expect rank-identical:")
            for problem in problems:
                print(f"  {problem}")
            print("\nThis build claimed to leave ranking alone and did not. "
                  "Do not draft from it until you know why.")
            return 1
        print("\nPASSED --expect rank-identical: every rank, VOR and Adj PPG "
              "is unchanged. Only the columns that were supposed to move did.")
    return 0


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("old")
    parser.add_argument("new")
    parser.add_argument("--top", type=int, default=60,
                        help="position-mix summary depth")
    parser.add_argument("--focus", type=int, default=200,
                        help="only list rank moves for players inside this "
                             "rank on either board; deeper moves are counted "
                             "but not printed")
    parser.add_argument("--expect", choices=["any", "rank-identical"],
                        default="any",
                        help="'rank-identical' exits non-zero if any rank, "
                             "VOR or Adj PPG moved")
    args = parser.parse_args()
    sys.exit(compare(args.old, args.new, top=args.top, expect=args.expect,
                     focus=args.focus))


if __name__ == "__main__":
    main()
