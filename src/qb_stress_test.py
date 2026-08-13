"""
Sweeps `expected_drafted.QB` and reports which players survive every
scenario. A 32-team superflex tool.

WHY THIS EXISTS
---------------
The 32-team board's replacement level comes from an `expected_drafted`
block typed in from ONE mock draft in which 31 of 32 teams autodrafted.
QB59 was the most load-bearing number on that board and it was a sample
of one. Phase 13.6 added a second mock and averaged the two, which moves
the shipped count to QB62 -- still the same side of the cliff below, and
still only two drafts. The sweep is why the number does not have to be
exactly right.

The board turns out to be BIMODAL in it. Sweeping 45 to 70:

    QB <= 49    8 quarterbacks in the top 60
    QB >= 55   18 quarterbacks in the top 60

with the transition between 49 and 50. So the question is not "is 59
exactly right" but "which side of the cliff is the room on," and one
mock's count is being asked to answer it.

WHAT IT PRODUCES, AND WHY IT IS A WORST-CASE RANK
-------------------------------------------------
For every player, the rank he holds under each scenario. `Worst rank` is
where he lands if the QB assumption breaks against him. Draft off that
column and you cannot be wrong by more than the swing -- it is the rank
that survives being wrong about the one thing this board is least sure
of.

Non-QB players barely move: the whole instability is concentrated in the
QB6-QB16 tier, which swings 40-80 places. Elite quarterbacks are safe in
every scenario and so is every running back.

THIS MUST BE REGENERATED WHENEVER THE BOARD IS (Aug 12). The first
version was built from v13 and was still on disk after the v15 refresh
had moved 106 players' Adj PPG -- a worst-case rank computed from
superseded projections, sitting next to a current board, four hours
before a draft. Hence a module rather than a one-off script.

USAGE
-----
    python -m src.qb_stress_test
    python -m src.qb_stress_test --board 2026_32Team_Board_v15.xlsx
"""

import argparse
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_BOARD = PROJECT_ROOT / "2026_32Team_Board_v17.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "QB_stress_test.xlsx"

HEADER_ROW = 7
# PHASE 13.6. Was QB59/RB92/WR145/TE56 = 352, off ONE 11-round mock.
# The league went to 12 rounds and a second mock exists, so these are now
# the average of the two (the 11-round one put on a 12-round scale first)
# and they sum to 384. Kept in sync with league_config_32team.json by
# hand -- if they disagree, this file is stress-testing a board that
# nobody shipped.
SHIPPED = {"QB": 62, "RB": 99, "WR": 162, "TE": 61}
SWEEP = list(range(45, 71))
REPORTED = [45, 49, 52, 55, 59, 62, 65, 70]
TOP = 60

FILLS = {"QB": "CFE2F3", "RB": "D9EAD3", "WR": "FCE5CD", "TE": "EAD1DC"}


def load_board(path):
    board = pd.read_excel(path, sheet_name="Draft Board", header=HEADER_ROW)
    board = board[board["Rank"].notna()].copy()
    board["Rank"] = board["Rank"].astype(int)
    return board


def redistribute(qb_count, total):
    """
    Move the QB delta onto RB/WR/TE in proportion to their observed shares.

    The total must stay fixed: it is teams x rounds minus the picks that
    go to positions this model does not rank, and it does not change
    because the room's taste in quarterbacks does. Shifting QB without
    compensating elsewhere would be sweeping the draft's LENGTH, which is
    a different and uninteresting question.
    """
    rest = total - qb_count
    others = {p: SHIPPED[p] for p in ("RB", "WR", "TE")}
    scale = sum(others.values())
    exact = {p: others[p] * rest / scale for p in others}
    counts = {p: int(np.floor(v)) for p, v in exact.items()}
    for position in sorted(exact, key=lambda p: -(exact[p] - counts[p]))[
        : rest - sum(counts.values())
    ]:
        counts[position] += 1
    return {"QB": qb_count, **counts}


def rank_under(board, counts):
    """Re-rank the whole board with these replacement ranks."""
    frames = []
    for position, rank in counts.items():
        pool = (board[board.Pos == position]
                .sort_values("Adj PPG", ascending=False)
                .reset_index(drop=True))
        replacement = pool.loc[min(rank - 1, len(pool) - 1), "Adj PPG"]
        pool = pool.assign(vor=pool["Adj PPG"] - replacement)
        frames.append(pool[["Player", "Pos", "vor"]])
    ranked = (pd.concat(frames)
              .sort_values("vor", ascending=False)
              .reset_index(drop=True))
    ranked["rank"] = ranked.index + 1
    return ranked.set_index("Player")["rank"]


def build(board_path=DEFAULT_BOARD, output_path=DEFAULT_OUTPUT):
    board = load_board(board_path)
    total = sum(SHIPPED.values())
    print(f"Board: {Path(board_path).name}  ({len(board)} players)")
    print(f"Sweeping expected_drafted.QB over {SWEEP[0]}-{SWEEP[-1]}, "
          f"total held at {total} skill picks.\n")

    ranks = pd.DataFrame({qb: rank_under(board, redistribute(qb, total))
                          for qb in SWEEP})
    meta = board.set_index("Player")[["Pos", "Adj PPG", "Team", "Bye", "ADP (Ovr)"]]
    table = ranks.join(meta)
    table["best"] = ranks.min(axis=1)
    table["worst"] = ranks.max(axis=1)
    table = table.sort_values("worst").reset_index()

    print(f"{'QB drafted':>10}  top-60 mix")
    for qb in REPORTED:
        top = table.nsmallest(TOP, qb)
        mix = top.Pos.value_counts()
        print(f"{qb:>10}  " + "  ".join(
            f"{p} {int(mix.get(p, 0)):2d}" for p in ("QB", "RB", "WR", "TE")))

    robust = table[table.worst <= TOP]
    print(f"\nROBUST: {len(robust)} players are top-{TOP} under EVERY scenario "
          f"-- " + ", ".join(
              f"{p} {int(c)}" for p, c in robust.Pos.value_counts().items()))

    _write(table, robust, output_path, board_path)
    return table


def _write(table, robust, output_path, board_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Robust Board"
    thin = Side(style="thin", color="BFBFBF")

    sheet["A1"] = "32-Team Superflex — QB replacement-level stress test"
    sheet["A1"].font = Font(name="Arial", size=14, bold=True)
    notes = [
        f"Source board: {Path(board_path).name}. Regenerate this file whenever "
        f"that board is rebuilt — a worst-case rank from a superseded board is "
        f"worse than none.",
        "Method: sweep expected_drafted.QB from 45 to 70, redistributing the "
        "difference across RB/WR/TE by observed share so the total stays at "
        f"{sum(SHIPPED.values())} skill picks. Adj PPG never changes; only replacement level, and so rank.",
        "The board is BIMODAL: at or below QB49, 8 quarterbacks make the top 60; "
        f"at or above QB55, 18 do. The measured value ({SHIPPED['QB']}) sits on the "
        "stable side.",
        "USE: 'Worst rank' is where a player lands if the QB assumption breaks "
        "against him. Draft off it and you cannot be wrong by more than the swing. "
        "Sorted by it.",
    ]
    row = 2
    for note in notes:
        cell = sheet.cell(row, 1, note)
        cell.font = Font(name="Arial", size=9, italic=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
        sheet.row_dimensions[row].height = 28
        row += 1

    header = row + 1
    columns = (["Worst rank", "Rank @ QB59 (shipped)", "Best rank", "Swing",
                "Player", "Pos", "Team", "Bye", "Adj PPG", "ADP (Ovr)"]
               + [f"@QB{q}" for q in REPORTED])
    for index, label in enumerate(columns, 1):
        cell = sheet.cell(header, index, label)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="434343")
        cell.alignment = Alignment(wrap_text=True, horizontal="center",
                                   vertical="center")
    sheet.row_dimensions[header].height = 32

    for offset, record in table.iterrows():
        line = header + 1 + offset
        values = ([record["worst"], record[59], record["best"], None,
                   record["Player"], record["Pos"], record["Team"],
                   record["Bye"], record["Adj PPG"], record["ADP (Ovr)"]]
                  + [record[q] for q in REPORTED])
        for index, value in enumerate(values, 1):
            cell = sheet.cell(line, index, value)
            cell.font = Font(name="Arial", size=10)
            cell.border = Border(bottom=thin)
            if index == 9:
                cell.number_format = "0.00"
        sheet.cell(line, 4, f"=A{line}-C{line}").font = Font(name="Arial", size=10)
        fill = FILLS.get(record["Pos"])
        if fill:
            for index in range(1, len(columns) + 1):
                sheet.cell(line, index).fill = PatternFill("solid", fgColor=fill)
        if record["worst"] <= TOP:
            sheet.cell(line, 5).font = Font(name="Arial", size=10, bold=True)

    for index, width in enumerate([11, 13, 10, 8, 24, 6, 7, 6, 10, 11]
                                  + [8] * len(REPORTED), 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = sheet.cell(header + 1, 5)
    sheet.auto_filter.ref = (
        f"A{header}:{get_column_letter(len(columns))}{header + len(table)}"
    )

    workbook.save(output_path)
    print(f"\nWrote {Path(output_path).name} "
          f"({len(table)} players, {len(robust)} bolded as robust)")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--board", default=str(DEFAULT_BOARD))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()
    build(args.board, args.output)


if __name__ == "__main__":
    main()
