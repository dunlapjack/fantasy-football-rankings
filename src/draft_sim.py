"""
Phase 14 -- the draft-day strategy bakeoff.

WHY THIS FILE EXISTS
--------------------
Every phase so far has answered "who is best." None has answered "who should
I take, given who I already have." Those are different questions and the
board only answers the first one.

The board sorts by VOR, and VOR in `build_board.py` is computed ONCE against
an empty draft: a player's points per game minus the points per game of the
last player drafted at his position. That number never moves. It cannot,
because it is a property of the player and the league, not of your roster.

So on the clock in round 8, holding four running backs, the board still
prices RB5 at whatever he was worth in round 1 -- even though he will start
zero games for you. Best-available on a static-VOR board is not
best-available *to you*. The same mechanism recommends a second quarterback
in a one-quarterback league: QB2's VOR is positive, and positive sorts above
the WR whose VOR is smaller but who would actually enter your lineup.

That is the whole diagnosis. This file measures the cost.

WHAT IT DOES
------------
Monte Carlo. Simulate the draft, simulate the season, count the points.

  1. Opponents draft by ADP with noise, subject to their own roster caps.
  2. My team drafts by one of several POLICIES (see POLICIES below).
  3. The resulting roster is scored by playing out the league's regular
     season week by week: byes are deterministic, injuries are stochastic,
     the lineup is filled optimally from whoever is available, and any slot
     that cannot be filled falls back to the best undrafted player at that
     position -- because in a real league you would stream him.
  4. Repeat, and compare policies by mean starting-lineup points.

The week-by-week loop is not decoration. If you score a roster by its
starters alone, bench players are worth exactly zero and the simulator will
always tell you to fill your starting lineup and then draft nothing -- a
degenerate answer produced by the scoring function, not by football. Byes
and injuries are what make the 4th running back worth anything, so they have
to be in the season model or the experiment is rigged before it starts.

WHAT IS MEASURED VS WHAT IS ASSUMED
-----------------------------------
Measured (comes from the board / the data):
  - every player's projected points per game, and his bye week
  - ADP, and its standard deviation, which sets opponent noise
  - replacement level, per league, from `build_board.py`

Assumed (stated here so it can be argued with):
  - opponents draft ADP + Gaussian noise with positional caps. Real drafters
    are worse than this in ways that would help me, so this is conservative.
  - injuries are a two-state Markov chain: healthy -> hurt with probability
    h, hurt -> healthy with probability 1/3 (so absences average 3 weeks).
    h is solved per position to hit a target season miss-rate.
  - weekly scoring noise is lognormal-ish with a per-position coefficient of
    variation, used ONLY for the head-to-head win-rate metric.

The injury and noise assumptions are the load-bearing ones, so `--sweep`
re-runs the whole bakeoff at three injury levels. If the ranking of policies
is stable across them, the conclusion does not depend on the assumption. If
it is not stable, that is the finding and it gets reported as one.

Standing project rule, obeyed here: ADP never enters a projection. It sets
opponent behavior and replacement level, nothing else.

USAGE
-----
    python -m src.draft_sim                      # all four leagues
    python -m src.draft_sim --league 12team
    python -m src.draft_sim --sims 1000 --sweep
"""

import argparse
import json
import math
import random
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# ---------------------------------------------------------------------
# League registry. Draft slot is 1-based. `keepers` are players already on
# MY roster; `gone` are players kept by somebody else and therefore absent
# from the pool.
# ---------------------------------------------------------------------
LEAGUES = {
    "12team": {
        "config": "league_config_12team.json",
        "board": "2026_12Team_Board_v17.xlsx",
        "slot": 3,
        "keepers": ["Chris Olave"],
        "keeper_rounds_forfeited": [8],
        "gone": ["Ja'Marr Chase", "Jonathan Taylor", "Javonte Williams",
                 "George Pickens"],
        "predrafted": {},
    },
    "8team": {
        "config": "league_config_8team.json",
        "board": "2026_8Team_Board_v17.xlsx",
        "slot": None,          # unknown -- average over every slot
        "keepers": [], "keeper_rounds_forfeited": [], "gone": [],
        "predrafted": {},
    },
    "6team": {
        "config": "league_config_6team.json",
        "board": "2026_6Team_Board_v17.xlsx",
        "slot": 5,
        "keepers": [], "keeper_rounds_forfeited": [], "gone": [],
        "predrafted": {},
    },
    "32team": {
        "config": "league_config_32team.json",
        "board": "2026_32Team_Board_v17.xlsx",
        "slot": 4,
        "keepers": [], "keeper_rounds_forfeited": [],
        "gone": [],
        # Round 1 is already spent. This is not a keeper -- it is a pick
        # that has happened, so it consumes the round.
        "predrafted": {1: "Puka Nacua"},
    },
}

MODELED = ("QB", "RB", "WR", "TE")

# Slots this model does not rank. Every team spends its last two picks here
# and the choice is uncorrelated with skill-position strategy, so they are
# removed from the draft entirely rather than simulated. Removing them is
# not the same as ignoring them: a 16-round draft with K and DST is a
# 14-round skill draft, and pretending otherwise would hand every policy two
# extra bench picks it does not really have.
UNMODELED_SLOTS = ("K", "DST")

# Season miss-rate targets: the share of the regular season an average
# player at each position misses. Higher at RB, which is the position whose
# bench depth is most in question, so getting this wrong runs AGAINST the
# conclusion this file reaches rather than toward it.
BASE_MISS_RATE = {"QB": 0.09, "RB": 0.17, "WR": 0.12, "TE": 0.14}
RECOVERY_PROB = 1.0 / 3.0      # absences average three weeks

# Weekly scoring noise, used only for the head-to-head win metric. Points
# totals are the primary result and do not touch this.
WEEKLY_CV = {"QB": 0.34, "RB": 0.48, "WR": 0.52, "TE": 0.56}

# How deep past the last drafted player at a position the "streamable"
# waiver option sits. 3rd-best undrafted, not best: other teams are picking
# off waivers too, and assuming you always land the top one is the standard
# way these simulators quietly inflate thin rosters.
WAIVER_DEPTH = 3

NFL_WEEKS = 18


# ---------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------
class Player:
    __slots__ = ("name", "pos", "ppg", "adp", "adp_sd", "bye", "rank",
                 "team", "exp_games", "idx")

    def __init__(self, name, pos, ppg, adp, adp_sd, bye, rank, team,
                 exp_games, idx):
        self.name, self.pos, self.ppg = name, pos, ppg
        self.adp, self.adp_sd, self.bye = adp, adp_sd, bye
        self.rank, self.team, self.exp_games, self.idx = rank, team, exp_games, idx

    def __repr__(self):
        return f"{self.name} ({self.pos}, {self.ppg:.1f})"


def load_board(board_path, features_path=None):
    """
    Reads the shipped Excel board rather than re-deriving it.

    The board is the project's deliverable and it has already been through
    the holdout gate, the playing-time gate, position overrides, per-league
    rescoring and the mock-ADP path for the 32-team league. Recomputing any
    of that here would create a second source of truth that could disagree
    with the spreadsheet Jack actually drafts from -- and the version that
    disagreed would be this one.
    """
    wb = load_workbook(board_path, read_only=True, data_only=True)
    ws = wb["Draft Board"]

    # FIND the header row; do not assume it. It was row 8 through v17 and
    # moved to row 12 in v17.1 when the Phase 14 notes block grew, and a
    # hardcoded 8 would not have failed loudly -- it would have read the
    # notes paragraph as column labels, found no "Adj PPG", and raised
    # something misleading. `compare_boards.py` already scans for it, which
    # is the precedent worth copying.
    header, header_row = None, None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40,
                                         values_only=True), start=1):
        if row and "Player" in row and "Adj PPG" in row:
            header, header_row = list(row), i
            break
    if header is None:
        raise ValueError(f"{board_path}: no header row found in the first 40 "
                         f"rows (looked for 'Player' and 'Adj PPG')")

    rows = [r for r in ws.iter_rows(min_row=header_row + 1, values_only=True)
            if r and r[0] is not None]
    wb.close()

    col = {label: i for i, label in enumerate(header) if label}
    for required in ("Pos", "Player", "Adj PPG", "ADP (Ovr)", "Bye",
                     "Rank", "Has ADP", "Draft Target"):
        if required not in col:
            raise KeyError(f"{board_path}: board is missing column {required!r}")

    # adp_stdev is not on the board but is in player_features.csv, and it is
    # what makes opponent noise player-specific instead of a flat guess.
    sd_by_name = {}
    if features_path and Path(features_path).exists():
        import csv
        with open(features_path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                for key in ("adp_stdev", "adp_stdev_2qb"):
                    try:
                        v = float(row.get(key) or "")
                    except ValueError:
                        continue
                    if v > 0:
                        sd_by_name.setdefault(row["player_name"], v)

    players, idx = [], 0
    for r in rows:
        pos = r[col["Pos"]]
        if pos not in MODELED:
            continue
        target = str(r[col["Draft Target"]] or "")
        if "DO NOT DRAFT" in target:
            continue                       # out for the season
        ppg = r[col["Adj PPG"]]
        if ppg is None:
            continue
        has_adp = bool(r[col["Has ADP"]])
        adp = r[col["ADP (Ovr)"]]
        if has_adp and adp is None:
            continue
        name = r[col["Player"]]
        players.append(Player(
            name=name, pos=pos, ppg=float(ppg),
            adp=float(adp) if has_adp else None,
            adp_sd=float(sd_by_name.get(name, 8.0)),
            bye=int(r[col["Bye"]]) if r[col["Bye"]] else 0,
            rank=int(r[col["Rank"]]), team=r[col["Team"]],
            exp_games=float(r[col["Exp Gm"]] or 0), idx=idx))
        idx += 1

    # Players with no ADP get a synthetic one that places them behind every
    # ADP-bearing player, in board-rank order. This mirrors what
    # build_board.py already does -- it hard-caps them below the market --
    # and it is not cosmetic here: the 32-team league has 384 picks against
    # 306 players with a real ADP, so without this tail the draft runs the
    # pool dry, the last rounds pick from nothing, and every waiver level
    # collapses to zero. That failure inflates depth-heavy strategies,
    # which is the exact thing under test.
    real = [p for p in players if p.adp is not None]
    tail = sorted((p for p in players if p.adp is None), key=lambda p: p.rank)
    last = max((p.adp for p in real), default=0.0)
    for i, p in enumerate(tail, start=1):
        p.adp = last + i
        p.adp_sd = 12.0
    return players


def load_config(path):
    with open(path) as f:
        return json.load(f)


def lineup_spec(config):
    """
    (dedicated slots, flex count, superflex count) for the modeled positions.

    K and DST are dropped -- see UNMODELED_SLOTS.
    """
    slots = config["roster_slots"]
    dedicated = {p: int(slots.get(p, 0)) for p in MODELED}
    return dedicated, int(slots.get("FLEX", 0)), int(slots.get("SUPERFLEX", 0))


def skill_rounds(config):
    """Draft length in rounds that actually spend a skill pick."""
    unmodeled = sum(1 for s in UNMODELED_SLOTS if config["roster_slots"].get(s))
    return int(config["total_rounds"]) - unmodeled


# ---------------------------------------------------------------------
# Lineup optimisation
# ---------------------------------------------------------------------
def best_lineup(available, dedicated, flex, superflex, waiver):
    """
    Fills the starting lineup from `available` and returns its total PPG.

    Greedy, most-constrained-first, which is exactly optimal here because
    the eligibility sets are nested: a dedicated slot accepts one position,
    FLEX accepts RB/WR/TE, SUPERFLEX accepts all four. With nested
    eligibility the greedy assignment is provably the maximum, so there is
    no need for a matching algorithm.

    Any slot that cannot be filled falls back to `waiver[pos]` rather than
    scoring zero. A zero would say "you fielded nobody," which no league
    makes you do -- you pick up whoever is on the wire. Modelling that as
    zero is the single easiest way to make depth look more valuable than it
    is, so it is worth being explicit that this is a floor, not an absence.
    """
    by_pos = defaultdict(list)
    for p in available:
        by_pos[p.pos].append(p)
    for pos in by_pos:
        by_pos[pos].sort(key=lambda p: -p.ppg)

    total, used = 0.0, {p: 0 for p in MODELED}
    for pos, count in dedicated.items():
        for _ in range(count):
            pool = by_pos.get(pos, [])
            if used[pos] < len(pool):
                total += pool[used[pos]].ppg
                used[pos] += 1
            else:
                total += waiver[pos]

    for eligible, count in ((("RB", "WR", "TE"), flex),
                            (("QB", "RB", "WR", "TE"), superflex)):
        for _ in range(count):
            best, best_pos = None, None
            for pos in eligible:
                pool = by_pos.get(pos, [])
                if used[pos] < len(pool):
                    cand = pool[used[pos]]
                    if best is None or cand.ppg > best.ppg:
                        best, best_pos = cand, pos
            if best is not None:
                total += best.ppg
                used[best_pos] += 1
            else:
                total += max(waiver[p] for p in eligible)
    return total


# ---------------------------------------------------------------------
# Draft policies
#
# Each takes (available, roster, ctx) and returns the chosen Player.
# `ctx` carries the league shape, how many picks I have left, and how many
# picks pass before my next turn.
# ---------------------------------------------------------------------
def policy_adp(available, roster, ctx):
    """Control: draft the consensus board. What a normal manager does."""
    return min(available, key=lambda p: p.adp)


def policy_board(available, roster, ctx):
    """
    Pure best-available off the v17 board. This is the behaviour that
    produces five running backs and two quarterbacks before a second
    receiver, and it is the thing every other policy is measured against.
    """
    return min(available, key=lambda p: p.rank)


def _slots_open(roster, ctx):
    """Starting slots not yet covered, as a set of positions worth filling."""
    dedicated, flex, superflex = ctx["dedicated"], ctx["flex"], ctx["superflex"]
    have = defaultdict(int)
    for p in roster:
        have[p.pos] += 1
    need = set()
    spare = 0
    for pos, count in dedicated.items():
        if have[pos] < count:
            need.add(pos)
        else:
            spare += have[pos] - count
    if spare < flex + superflex:
        need.update(("RB", "WR", "TE"))
        if superflex:
            need.add("QB")
    return need


def policy_starters_first(available, roster, ctx):
    """
    Fill every starting slot before taking anyone else, then revert to the
    board. The strict version of "fill your lineup first."
    """
    need = _slots_open(roster, ctx)
    pool = [p for p in available if p.pos in need] if need else available
    return min(pool or available, key=lambda p: p.rank)


def policy_caps(available, roster, ctx):
    """
    Best-available, but never exceed a sane count at a position. The
    cheapest possible fix and the one worth beating: if a rule this crude
    captures most of the gain, nothing more complicated is justified.
    """
    have = defaultdict(int)
    for p in roster:
        have[p.pos] += 1
    caps = dict(ctx["caps"])
    # Late in the draft the quarterback cap relaxes -- a backup QB in the
    # last two rounds is a real thing to do, just not in round 6.
    if ctx["picks_left"] <= 3:
        caps["QB"] = caps["QB"] + 1
        caps["TE"] = caps["TE"] + 1
    need = _slots_open(roster, ctx)
    pool = [p for p in available
            if have[p.pos] < caps.get(p.pos, 99) or p.pos in need]
    return min(pool or available, key=lambda p: p.rank)


def _next_available_at(available_by_adp, gap):
    """
    Best remaining player at each position after the next `gap` picks, if
    those picks follow ADP order.

    This is the scarcity estimate. It is deliberately deterministic -- ADP
    order with no noise -- because the point is the SHAPE of the drop-off at
    each position, and averaging noisy draws would give the same shape at
    several times the cost.
    """
    taken = set()
    for p in available_by_adp[:gap]:
        taken.add(p.idx)
    best = {}
    for p in available_by_adp:
        if p.idx in taken:
            continue
        if p.pos not in best or p.ppg > best[p.pos].ppg:
            best[p.pos] = p
    return best


def policy_vona(available, roster, ctx):
    """
    Value Over Next Available. Rank by how much better a player is than the
    best man at his own position expected to survive until my next pick.

    This prices scarcity correctly -- it is why you take the last elite
    tight end early -- but it is blind to your roster. VONA will happily
    hand you a fifth running back if running backs are falling off a cliff,
    because it never asks whether he would play. Included precisely to
    separate the two effects: if VONA alone closes the gap, the problem was
    scarcity; if it does not, the problem was roster construction.
    """
    by_adp = sorted(available, key=lambda p: p.adp)
    fallback = _next_available_at(by_adp, ctx["gap"])
    best, best_key = None, None
    for p in available:
        alt = fallback.get(p.pos)
        val = p.ppg - (alt.ppg if alt is not None else 0.0)
        key = (-round(val, 6), p.rank)
        if best_key is None or key < best_key:
            best, best_key = p, key
    return best


def availability_scenarios(config, count=3, seed=0):
    """
    A few pre-drawn injury patterns, used by `policy_marginal` to value
    bench depth while it is on the clock.

    WHY THIS IS NOT OPTIONAL. Score a candidate roster on byes alone and
    every bench player is worth the same thing: nothing, except in the one
    week his position-mate is off. That produced a decision function where
    the top six candidates all tied at 0.00 and the pick fell to whatever
    the sort happened to put first -- so the policy was choosing at random
    and the resulting rosters were nonsense (six quarterbacks) for a reason
    that had nothing to do with football.

    Depth is insurance. Insurance has no value in a world with no
    accidents, so the world the policy reasons about has to have some.

    The scenarios are drawn ONCE per league and reused for every candidate
    and every policy, which makes them common random numbers: they cannot
    favour one pick over another, they only supply the texture that breaks
    the ties honestly.
    """
    weeks = list(range(config["regular_season_weeks"][0],
                       config["regular_season_weeks"][1] + 1))
    rng = random.Random(seed)
    out = []
    for _ in range(count):
        # A scenario is a set of (position, week) absences expressed as a
        # per-week draw the policy applies to whichever players it holds.
        # Storing it per position rather than per player keeps it valid as
        # the roster changes mid-draft, which is the whole point.
        pattern = {}
        for pos in MODELED:
            hazard = injury_hazard(BASE_MISS_RATE[pos])
            for depth in range(8):
                state, marks = False, []
                for _w in weeks:
                    if state:
                        state = rng.random() >= RECOVERY_PROB
                    else:
                        state = rng.random() < hazard
                    marks.append(state)
                # Keyed by (position, depth) rather than by player, so the
                # pattern stays meaningful as the roster fills up and the
                # k-th best receiver becomes a different man. Independent
                # per depth slot -- keying by position alone would knock
                # out every receiver in the same week, which would make
                # depth look useless rather than valuable.
                pattern[(pos, depth)] = marks
        out.append((weeks, pattern))
    return out


def _lineup_gain_quick(roster, p, dedicated, flex, superflex, waiver):
    """Cheap no-bye, no-injury lineup improvement. Used only inside the rollout."""
    before = best_lineup(roster, dedicated, flex, superflex, waiver)
    return best_lineup(roster + [p], dedicated, flex, superflex, waiver) - before


def policy_marginal(available, roster, ctx):
    """
    One-ply lookahead with a greedy completion of the rest of the draft.

    For each candidate: put him on the roster, then FINISH the roster by
    walking my remaining picks, at each one taking the best player expected
    to still be available at that pick. Score the completed roster over the
    injury scenarios. Take the candidate whose finished team scores highest.

    WHY THE SIMPLER VERSION DID NOT WORK, because it is the interesting part.

    The obvious rule is greedy: value a player by what he adds to my lineup
    now, minus what the best man at his position would add if I waited one
    pick. That rule is correct in its parts and useless in practice, and the
    reason is worth writing down.

    Once my starting lineup is covered, EVERY candidate's value collapses to
    approximately zero -- nobody improves a lineup that is already full, and
    the wait-one-pick term cancels most of what is left. Dozens of players
    tie at 0.00 and the pick falls to the tiebreak. Whatever the tiebreak is,
    that is now the real policy. Ours was board rank, so the "sophisticated"
    policy quietly degenerated into pure best-available -- the exact
    behaviour it was written to beat -- and drafted five quarterbacks while
    appearing to reason about lineups.

    A one-pick horizon is what causes it. Skipping a receiver in round 6 is
    only cheap if I take one in round 7, and the greedy rule assumes I will
    without ever checking. Rolling the draft forward to the end forces the
    check: a roster that ends with two receivers scores badly, and that
    shows up as a lower number on the candidate that led there, in round 6,
    where the decision actually gets made.

    There is still no tuning constant anywhere in this. The rollout is
    greedy by lineup improvement, the availability estimate is ADP order,
    and everything else is the projections and the schedule.
    """
    dedicated, flex, superflex = ctx["dedicated"], ctx["flex"], ctx["superflex"]
    waiver, scenarios = ctx["waiver"], ctx["scenarios"]
    by_adp = sorted(available, key=lambda p: p.adp)

    def season(players):
        order = defaultdict(list)
        for p in sorted(players, key=lambda p: -p.ppg):
            order[p.pos].append(p)
        depth = {}
        for pos, group in order.items():
            for i, p in enumerate(group):
                depth[p.idx] = (pos, min(i, 7))
        total = 0.0
        for weeks, pattern in scenarios:
            for wi, w in enumerate(weeks):
                total += best_lineup(
                    [p for p in players
                     if p.bye != w and not pattern[depth[p.idx]][wi]],
                    dedicated, flex, superflex, waiver)
        return total / len(scenarios)

    # Candidate set measured in POINTS, the same unit the lineup is scored
    # in -- not board rank, which is VOR order and a different ordering.
    seen, candidates = set(), []
    by_pos = defaultdict(list)
    for p in available:
        by_pos[p.pos].append(p)
    for pos, group in by_pos.items():
        for p in sorted(group, key=lambda q: -q.ppg)[:5]:
            if p.idx not in seen:
                seen.add(p.idx)
                candidates.append(p)
    for p in sorted(available, key=lambda q: q.rank)[:10]:
        if p.idx not in seen:
            seen.add(p.idx)
            candidates.append(p)

    gaps = ctx["future_gaps"]          # picks that pass before each of my turns
    best, best_key = None, None
    for cand in candidates:
        roster2 = list(roster) + [cand]
        used = {cand.idx}
        for g in gaps:
            pool = [p for p in by_adp[g:] if p.idx not in used][:45]
            if not pool:
                break
            nxt = max(pool, key=lambda p: (
                _lineup_gain_quick(roster2, p, dedicated, flex, superflex, waiver),
                p.ppg))
            roster2.append(nxt)
            used.add(nxt.idx)
        key = (-round(season(roster2), 6), cand.rank)
        if best_key is None or key < best_key:
            best, best_key = cand, key
    return best


POLICIES = {
    "adp": policy_adp,
    "board": policy_board,
    "starters_first": policy_starters_first,
    "caps": policy_caps,
    "vona": policy_vona,
    "marginal": policy_marginal,
}


# ---------------------------------------------------------------------
# The draft
# ---------------------------------------------------------------------
def opponent_pick(available, roster, caps, need, rng):
    """
    Opponents take the best player by ADP-plus-noise, skipping positions
    they have filled, and are forced onto need if they are running out of
    picks.

    The noise is each player's own `adp_stdev`, so consensus players go
    close to their ADP and divisive ones scatter -- which is the realistic
    thing and also the conservative one. A perfectly-ADP opponent field
    would be easier to exploit than a real room, and would flatter every
    policy here equally, but a NOISELESS field would make the "next
    available at my turn" estimates suspiciously accurate and inflate the
    two policies that use them.
    """
    have = defaultdict(int)
    for p in roster:
        have[p.pos] += 1
    pool = [p for p in available
            if have[p.pos] < caps.get(p.pos, 99) or p.pos in need]
    if need:
        forced = [p for p in pool if p.pos in need]
        if forced:
            pool = forced
    if not pool:
        pool = available
    pool = sorted(pool, key=lambda p: p.adp)[:40]
    return min(pool, key=lambda p: p.adp + rng.gauss(0, max(p.adp_sd, 2.0)))


def snake_order(teams, rounds):
    order = []
    for r in range(rounds):
        seq = range(teams) if r % 2 == 0 else range(teams - 1, -1, -1)
        order.extend((r, t) for t in seq)
    return order


def simulate_draft(players, config, league, policy_name, my_slot, rng,
                   scan=18):
    """Runs one full draft and returns my roster."""
    teams = int(config["num_teams"])
    rounds = skill_rounds(config)
    dedicated, flex, superflex = lineup_spec(config)
    weeks = list(range(config["regular_season_weeks"][0],
                       config["regular_season_weeks"][1] + 1))

    by_name = {p.name: p for p in players}
    available = {p.idx: p for p in players}

    rosters = [[] for _ in range(teams)]
    me = my_slot - 1

    # My keepers: on my roster, out of the pool, and they cost me the
    # forfeited rounds.
    for name in league["keepers"]:
        p = by_name.get(name)
        if p is None:
            raise KeyError(f"keeper {name!r} not on the board")
        rosters[me].append(p)
        available.pop(p.idx, None)

    # Everyone else's keepers. Each goes to a distinct opponent, who also
    # forfeits a pick, so the pick count stays honest. The round is drawn
    # rather than known -- an assumption, but a symmetric one: it changes
    # every policy's opponents identically, and the experiment is a
    # comparison between policies.
    opponents = [t for t in range(teams) if t != me]
    rng.shuffle(opponents)
    forfeited = defaultdict(set)
    for r in league["keeper_rounds_forfeited"]:
        forfeited[me].add(r - 1)
    for i, name in enumerate(league["gone"]):
        p = by_name.get(name)
        if p is None:
            continue
        owner = opponents[i % len(opponents)]
        rosters[owner].append(p)
        available.pop(p.idx, None)
        forfeited[owner].add(rng.randint(4, min(11, rounds - 1)))

    # Picks already made in a live draft.
    for rnd, name in league["predrafted"].items():
        p = by_name.get(name)
        if p is None:
            continue
        rosters[me].append(p)
        available.pop(p.idx, None)
        forfeited[me].add(rnd - 1)

    caps = ({"QB": 3, "RB": 6, "WR": 7, "TE": 2} if superflex
            else {"QB": 2, "RB": 6, "WR": 7, "TE": 2})
    waiver = waiver_levels(players, config)
    scenarios = availability_scenarios(config, count=3, seed=1)
    order = snake_order(teams, rounds)
    my_picks = [i for i, (r, t) in enumerate(order)
                if t == me and r not in forfeited[me]]
    policy = POLICIES[policy_name]

    for i, (rnd, team) in enumerate(order):
        if rnd in forfeited[team]:
            continue
        if not available:
            break
        avail = list(available.values())
        picks_left_for = sum(1 for j, (r2, t2) in enumerate(order)
                             if j > i and t2 == team and r2 not in forfeited[team])
        need = _slots_open(rosters[team],
                           {"dedicated": dedicated, "flex": flex,
                            "superflex": superflex})
        if team == me:
            future = [j for j in my_picks if j > i]
            gap = (future[0] - i) if future else 0
            # How many picks pass before each of my remaining turns, as
            # offsets into the ADP-ordered pool. This is what lets the
            # lookahead ask "who is realistically still there in round 9"
            # instead of assuming the board never moves.
            future_gaps = [j - i for j in future]
            ctx = {"dedicated": dedicated, "flex": flex, "superflex": superflex,
                   "waiver": waiver, "weeks": weeks, "gap": gap,
                   "picks_left": picks_left_for, "caps": caps, "scan": scan,
                   "scenarios": scenarios, "future_gaps": future_gaps}
            # Endgame guard applied to EVERY policy, including the naive
            # ones: with as many picks left as unfilled starting slots you
            # must fill them, or the roster is illegal. Without this the
            # naive policies would lose points to a technicality rather
            # than to their actual flaw, which would overstate the result.
            if need and picks_left_for < len(need):
                pool = [p for p in avail if p.pos in need]
                pick = min(pool or avail, key=lambda p: p.rank)
            else:
                pick = policy(avail, rosters[team], ctx)
        else:
            forced = need if picks_left_for < len(need) else set()
            pick = opponent_pick(avail, rosters[team], caps, forced, rng)
        rosters[team].append(pick)
        available.pop(pick.idx, None)

    return rosters[me], available


def waiver_levels(players, config):
    """
    Pre-draft ESTIMATE of the streamable option at each position: the
    WAIVER_DEPTH-th best player past the number expected to be drafted, in
    ADP order.

    Used by the policies while they are on the clock, because at that point
    the draft has not finished and the realized answer does not exist yet.
    Scoring uses `realized_waiver_levels` instead.
    """
    teams, rounds = int(config["num_teams"]), skill_rounds(config)
    picks = teams * rounds
    drafted = set(p.idx for p in sorted(players, key=lambda p: p.adp)[:picks])
    rest = [p for p in players if p.idx not in drafted]
    return _waiver_from(rest)


def realized_waiver_levels(leftover):
    """
    The streamable option as it actually turned out in THIS draft: the
    WAIVER_DEPTH-th best player nobody took.

    Worth the extra plumbing over the ADP estimate. The estimate assumes
    the room drafts exactly to ADP, and the whole simulation is built on
    the premise that it does not. In the 12-team league the two disagree at
    quarterback by several points per game, which is enough to change
    whether taking a second one looks defensible.
    """
    return _waiver_from(list(leftover))


def _waiver_from(pool):
    out = {}
    for pos in MODELED:
        rest = sorted((p for p in pool if p.pos == pos), key=lambda p: -p.ppg)
        out[pos] = rest[min(WAIVER_DEPTH - 1, len(rest) - 1)].ppg if rest else 0.0
    return out


# ---------------------------------------------------------------------
# The season
# ---------------------------------------------------------------------
def injury_hazard(miss_rate):
    """
    Weekly probability of an absence starting, solved so the long-run share
    of weeks missed equals `miss_rate` given RECOVERY_PROB.

    Two-state chain, stationary miss share = h / (h + r), so h = r*m/(1-m).
    """
    if miss_rate <= 0:
        return 0.0
    return RECOVERY_PROB * miss_rate / (1.0 - miss_rate)


def score_season(roster, config, waiver, rng, injury_scale=1.0,
                 opponents=None):
    """
    Plays the regular season and returns (starting-lineup points, weekly
    totals with noise).

    Absences are contiguous by construction -- the Markov chain has memory,
    so a player who is hurt in week 6 is likely hurt in week 7. That matters
    more than it sounds: bench value comes from absences OVERLAPPING with
    byes and with each other, and independent weekly coin flips would spread
    the damage evenly and understate exactly the depth this experiment is
    trying to price.
    """
    dedicated, flex, superflex = lineup_spec(config)
    weeks = list(range(config["regular_season_weeks"][0],
                       config["regular_season_weeks"][1] + 1))

    hurt = {p.idx: False for p in roster}
    hazard = {p.idx: injury_hazard(min(0.6, BASE_MISS_RATE[p.pos] * injury_scale))
              for p in roster}
    # Known preseason absence (PUP/NFI) from the board's Exp Gm column.
    forced_out = {}
    season_len = float(config["fantasy_season_length"])
    for p in roster:
        missed = max(0, round(season_len - p.exp_games))
        forced_out[p.idx] = set(weeks[:missed]) if missed else set()

    total, weekly = 0.0, []
    for w in weeks:
        for p in roster:
            if hurt[p.idx]:
                if rng.random() < RECOVERY_PROB:
                    hurt[p.idx] = False
            elif rng.random() < hazard[p.idx]:
                hurt[p.idx] = True
        avail = [p for p in roster
                 if p.bye != w and not hurt[p.idx] and w not in forced_out[p.idx]]
        pts = best_lineup(avail, dedicated, flex, superflex, waiver)
        total += pts
        weekly.append(pts)
    return total, weekly


def noisy_week(points, rng):
    """Applies weekly scoring variance for the head-to-head metric."""
    cv = 0.30
    return max(0.0, points * (1.0 + rng.gauss(0, cv)))


# ---------------------------------------------------------------------
# Experiment
# ---------------------------------------------------------------------
def run_league(key, sims=400, policies=None, injury_scale=1.0, seed=17,
               slots=None, quiet=False):
    league = LEAGUES[key]
    config = load_config(PROJECT_ROOT / league["config"])
    players = load_board(PROJECT_ROOT / league["board"],
                         PROJECT_ROOT / "data" / "player_features.csv")
    waiver = waiver_levels(players, config)
    policies = policies or list(POLICIES)
    if slots is None:
        slots = ([league["slot"]] if league["slot"]
                 else list(range(1, int(config["num_teams"]) + 1)))

    results = {}
    for name in policies:
        totals, rosters = [], []
        for s in range(sims):
            slot = slots[s % len(slots)]
            # Common random numbers: draft seed and season seed depend only
            # on the simulation index, never on the policy. So every policy
            # faces the SAME opponent draft order and the SAME injuries in
            # simulation s, and the difference between two policies is the
            # policies rather than the luck. This is what makes the paired
            # standard errors below small enough to separate them at a few
            # hundred sims instead of tens of thousands.
            rng = random.Random(seed * 100003 + s)
            roster, leftover = simulate_draft(players, config, league, name,
                                              slot, rng)
            rng2 = random.Random(seed * 7919 + s)
            pts, _ = score_season(roster, config,
                                  realized_waiver_levels(leftover.values()),
                                  rng2, injury_scale=injury_scale)
            totals.append(pts)
            rosters.append(roster)
        mean = sum(totals) / len(totals)
        var = sum((t - mean) ** 2 for t in totals) / max(1, len(totals) - 1)
        counts = defaultdict(float)
        for r in rosters:
            for p in r:
                counts[p.pos] += 1.0 / len(rosters)
        results[name] = {
            "mean": mean, "sd": math.sqrt(var),
            "se": math.sqrt(var / len(totals)),
            "pos": dict(counts), "totals": totals,
            "sample_roster": rosters[0],
        }
        if not quiet:
            shape = " ".join(f"{p}{counts[p]:.1f}" for p in MODELED)
            print(f"  {name:<16} {mean:8.1f} pts  (se {results[name]['se']:4.1f})   {shape}")

    # Paired comparison against pure best-available, which is the incumbent
    # behaviour and therefore the only baseline worth quoting a p-value
    # against.
    if "board" in results:
        base = results["board"]["totals"]
        for name, r in results.items():
            diffs = [a - b for a, b in zip(r["totals"], base)]
            m = sum(diffs) / len(diffs)
            v = sum((d - m) ** 2 for d in diffs) / max(1, len(diffs) - 1)
            r["delta"] = m
            r["delta_se"] = math.sqrt(v / len(diffs))
    return results, config, players


def head_to_head(key, sims=300, policies=None, seed=23):
    """
    How often each policy's team outscores a field of ADP-drafting teams,
    week by week, with scoring noise.

    Points are the primary metric; this exists because leagues are won by
    wins, and a policy that adds points in a way that does not convert into
    wins would be worth knowing about. It has an extra assumption baked in
    (the weekly CV) which is why it is reported second, not first.
    """
    league = LEAGUES[key]
    config = load_config(PROJECT_ROOT / league["config"])
    players = load_board(PROJECT_ROOT / league["board"],
                         PROJECT_ROOT / "data" / "player_features.csv")
    waiver = waiver_levels(players, config)
    policies = policies or list(POLICIES)
    slots = ([league["slot"]] if league["slot"]
             else list(range(1, int(config["num_teams"]) + 1)))

    out = {}
    for name in policies:
        wins = games = 0
        for s in range(sims):
            slot = slots[s % len(slots)]
            rng = random.Random(seed * 100003 + s)
            mine, left_a = simulate_draft(players, config, league, name, slot, rng)
            rng_b = random.Random(seed * 100003 + s + 5_000_000)
            theirs, left_b = simulate_draft(players, config, league, "adp",
                                            (slot % int(config["num_teams"])) + 1,
                                            rng_b)
            r1 = random.Random(seed * 31 + s)
            r2 = random.Random(seed * 37 + s)
            _, wa = score_season(mine, config,
                                 realized_waiver_levels(left_a.values()), r1)
            _, wb = score_season(theirs, config,
                                 realized_waiver_levels(left_b.values()), r2)
            rn = random.Random(seed * 41 + s)
            for a, b in zip(wa, wb):
                games += 1
                if noisy_week(a, rn) > noisy_week(b, rn):
                    wins += 1
        out[name] = wins / max(1, games)
    return out


def main():
    ap = argparse.ArgumentParser(description="Phase 14 draft strategy bakeoff.")
    ap.add_argument("--league", default=None, choices=list(LEAGUES))
    ap.add_argument("--sims", type=int, default=400)
    ap.add_argument("--seed", type=int, default=17)
    ap.add_argument("--sweep", action="store_true",
                    help="re-run at three injury levels")
    ap.add_argument("--h2h", action="store_true",
                    help="also compute head-to-head win rate vs an ADP field")
    ap.add_argument("--json", type=str, default=None)
    args = ap.parse_args()

    keys = [args.league] if args.league else list(LEAGUES)
    dump = {}
    for key in keys:
        cfg = load_config(PROJECT_ROOT / LEAGUES[key]["config"])
        print(f"\n=== {cfg['league_name']} ({key}) — {args.sims} drafts ===")
        res, _, _ = run_league(key, sims=args.sims, seed=args.seed)
        print(f"  {'':<16} {'vs board':>9} {'+/-':>6}  {'sigma':>6}")
        for name, r in sorted(res.items(), key=lambda kv: -kv[1]["mean"]):
            sig = r["delta"] / r["delta_se"] if r["delta_se"] else 0.0
            print(f"  {name:<16} {r['delta']:+9.1f} {r['delta_se']:6.1f}  {sig:6.1f}")
        dump[key] = {n: {k: v for k, v in r.items()
                         if k not in ("sample_roster", "totals")}
                     for n, r in res.items()}

        if args.sweep:
            print("  injury sensitivity (mean pts):")
            for scale, label in ((0.0, "none"), (1.0, "base"), (1.5, "high")):
                r2, _, _ = run_league(key, sims=max(150, args.sims // 3),
                                      injury_scale=scale, seed=args.seed,
                                      quiet=True)
                ranked = sorted(r2, key=lambda n: -r2[n]["mean"])
                print(f"    {label:<5} best={ranked[0]:<15} "
                      + "  ".join(f"{n}:{r2[n]['mean'] - r2['board']['mean']:+.0f}"
                                  for n in ranked))
                dump.setdefault(key, {}).setdefault("_sweep", {})[label] = {
                    n: r2[n]["mean"] for n in r2}

        if args.h2h:
            h = head_to_head(key, sims=max(100, args.sims // 4), seed=args.seed)
            print("  weekly win rate vs an ADP-drafting team:")
            for n, v in sorted(h.items(), key=lambda kv: -kv[1]):
                print(f"    {n:<16} {v * 100:5.1f}%")
            dump[key]["_h2h"] = h

    if args.json:
        with open(args.json, "w") as f:
            json.dump(dump, f, indent=2, default=str)
        print(f"\nwrote {args.json}")


if __name__ == "__main__":
    main()
